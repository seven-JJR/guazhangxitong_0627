import time,sys
import pymysql
from PyQt5.QtWidgets import QMessageBox, QInputDialog,QFileDialog
from PyQt5.QtWidgets import QMainWindow, QApplication, QDialog,QHBoxLayout ,QWidget
from PyQt5 import QtCore, QtGui, QtWidgets
from login_ui import login_ui_window
from interface_ui import interface_ui_window
from PyQt5.QtCore import Qt, QMargins, QPoint,QSettings,QDateTime,QTimer
import threading
from PyQt5.QtCore import pyqtSignal
import openpyxl
user_to_name={'D18030023':'蒋佳芮','D18030022':'大大'}
gonghao_all=['D18030023','D18030022']
user=''
#---------------登录界面----------------------
#---------------这一部分是UI逻辑分离固定格式----------------------
class loginwindow(QMainWindow,login_ui_window):
    def __init__(self):
        #super(login, self).__init__(parent)
        super().__init__()
        self.setupUi(self)
# ---------------这一部分是UI逻辑分离固定格式---------------------
        self.widget_3.hide()#隐藏注册页面
        self.pushButton_3.clicked.connect(self.switch_to_zhuceUI_xc)
        self.pushButton_2.clicked.connect(self.switch_to_loingUI_xc)
        self.pushButton_5.clicked.connect(self.zhuce)
        self.pushButton.clicked.connect(self.denglu)
        self.init_login_info()

# ---------------初始化登录信息，读取之前保存的data看是不是需要自动登录# ---------------
    def init_login_info(self):
        settings = QSettings("Mysoft","Lliy")
        username = settings.value("username")
        password = settings.value("password")
        remeberpassword = settings.value("remeberpassword")
        self.lineEdit.setText(username)#读取之前存的用户名运行app自动填入
        if remeberpassword == "true" or remeberpassword == True:#如果之前存的值是记住密码，就打开app就把密码自动填入
            self.checkBox.setChecked(True)
            self.lineEdit_2.setText(password)
# ---------------初始化登录信息，读取之前保存的data看是不是需要自动登录# ---------------

#---------------保存登录信息# ---------------
    def save_login_info(self):
        settings = QSettings("Mysoft", "Lliy")  # 方法2：使用注册表
        settings.setValue("username", self.lineEdit.text())
        settings.setValue("password", self.lineEdit_2.text())
        settings.setValue("remeberpassword", self.checkBox.isChecked())
# ---------------保存登录信息# ---------------

# ---------------这一部分是设置自定义窗口后，app不能拖动，重启可拖动的方式---------------------
    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.is_moving = True
            self.dragPos = event.globalPos() - self.pos()
            event.accept()
    def mouseMoveEvent(self, event):
        if self.is_moving:
            self.move(event.globalPos() - self.dragPos)
            event.accept()
# ---------------这一部分是设置自定义窗口后，app不能拖动，重启可拖动的方式---------------------

    def switch_to_zhuceUI_xc(self):
        tr_1 = threading.Thread(target=self.switch_to_zhuceUI)
        tr_1.setDaemon(True)
        tr_1.start()
    def switch_to_loingUI_xc(self):
        tr_1 = threading.Thread(target=self.switch_to_loingUI)
        tr_1.setDaemon(True)
        tr_1.start()
    def switch_to_zhuceUI(self):
       self.widget.hide()
       self.widget_3.show()
       self.label_3.setText('')
       self.lineEdit_3.clear()
       self.lineEdit_4.clear()
       self.lineEdit_6.clear()
    def switch_to_loingUI(self):
        self.widget_3.hide()
        self.widget.show()
        self.label_4.setText('')
    def denglu(self):
        global user
        self.save_login_info()
        try:
            db = pymysql.connect(host="localhost", user='root', password='jiang111', port=3306, database='lily', connect_timeout=3)
            cusor = db.cursor()
            sql_login_verify = f"SELECT * FROM users where Gong_Hao='{self.lineEdit.text()}' and Mi_Ma='{self.lineEdit_2.text()}'"
            cusor.execute(sql_login_verify)
            result=cusor.fetchone()
            if result is not None :#代表有找到正确的账户密码
                user = self.lineEdit.text()
                self.close()
                db.close()
                win_main_interface.show()
            else:
                self.label_4.setText('用户名或密码错误,请重新输入！')
                #self.label_4.repaint()
                db.close()
        except:
            self.label_4.setText('登录失败,网络连接异常！')
            #self.label_4.repaint()
    def zhuce(self):
        if len(self.lineEdit_3.text()) == 9 and self.lineEdit_4.text()==self.lineEdit_6.text():
            try:
                db=pymysql.connect(host="172.17.1.71", user='root', password='jiang111', port=3306, database='lily', connect_timeout=60)
                cusor=db.cursor()
                sql_insert_username = f"INSERT INTO users VALUES ('{self.lineEdit_3.text()}','{self.lineEdit_4.text()}')"
                cusor.execute(sql_insert_username)
                db.commit()
                self.label_3.setText('注册成功！')
                self.lineEdit_3.clear()
                self.lineEdit_4.clear()
                self.lineEdit_6.clear()
                db.close()
            except:
                self.label_3.setText('注册失败，请稍后再试')
        elif len(self.lineEdit_3.text()) != 9:
            self.label_3.setText('请输入正确的工号！')
        elif self.lineEdit_4.text()!=self.lineEdit_6.text():
            self.label_3.setText('密码不匹配！')
#---------------登录界面---------------------

#------------------------------主界面---------------------------------------#
# ---------------这一部分是UI逻辑分离固定格式---------------------
from datetime import datetime
class interfacewindow(QMainWindow, interface_ui_window):
    def __init__(self, parent=None):
        #super(login, self).__init__(parent)
        super().__init__()
        self.setupUi(self)
        self.wuzichazhao='wuzichazhao'#固定变量，数据库表格
        self.timer = QTimer()
        self.timer.timeout.connect(self.showTime)
        self.timer.start(1000)
        self.label_projectname_page4={0:self.label_92,1:self.label_93,2:self.label_94,3:self.label_95,4:self.label_98,5:self.label_99,6:self.label_166,7:self.label_175,8:self.label_178,9:self.label_188,10:self.label_194,11:self.label_209,12:self.label_226}
        self.jiechu_button_page4={0:self.pushButton_24,1:self.pushButton_47,2:self.pushButton_26,3:self.pushButton_48,4:self.pushButton_27,5:self.pushButton_49,6:self.pushButton_29,7:self.pushButton_50,8:self.pushButton_30,9:self.pushButton_51,10:self.pushButton_31,11:self.pushButton_52,12:self.pushButton_33}
        self.label_partno_page4={0:self.label_120,1:self.label_130,2:self.label_137,3:self.label_145,4:self.label_148,5:self.label_156,6:self.label_168,7:self.label_177,8:self.label_179,9:self.label_187,10:self.label_199,11:self.label_208,12:self.label_227}
        self.label_partname_page4={0:self.label_121,1:self.label_129,2:self.label_136,3:self.label_143,4:self.label_150,5:self.label_155,6:self.label_162,7:self.label_173,8:self.label_181,9:self.label_186,10:self.label_197,11:self.label_204,12:self.label_214}
        self.label_barcode_page4 = {0: self.label_122, 1: self.label_131, 2: self.label_138, 3: self.label_141, 4: self.label_149, 5: self.label_160, 6: self.label_164, 7: self.label_170, 8: self.label_184, 9: self.label_191, 10: self.label_196, 11: self.label_207, 12: self.label_225}
        self.label_satus_page4 = {0: self.label_123, 1: self.label_127, 2: self.label_134, 3: self.label_144, 4: self.label_153, 5: self.label_157, 6: self.label_165, 7: self.label_171, 8: self.label_185, 9: self.label_192, 10: self.label_200, 11: self.label_203, 12: self.label_213}
        self.label_projectname_page1 = {0: self.label_100, 1: self.label_518, 2: self.label_525, 3: self.label_531, 4: self.label_535, 5: self.label_544, 6: self.label_549, 7: self.label_772, 8: self.label_779, 9: self.label_782, 10: self.label_788, 11: self.label_796, 12: self.label_804}
        self.guihuan_button_page1 = {0: self.pushButton_13, 1: self.pushButton_14, 2: self.pushButton_15, 3: self.pushButton_16, 4: self.pushButton_19, 5: self.pushButton_18, 6: self.pushButton_17, 7: self.pushButton_20, 8: self.pushButton_21, 9: self.pushButton_200, 10: self.pushButton_201, 11: self.pushButton_202, 12: self.pushButton_203}
        self.label_partno_page1={0:self.label_125,1:self.label_514,2:self.label_521,3:self.label_534,4:self.label_541,5:self.label_547,6:self.label_550,7:self.label_766,8:self.label_774,9:self.label_781,10:self.label_791,11:self.label_800,12:self.label_806}
        self.label_partname_page1={0:self.label_126,1:self.label_519,2:self.label_526,3:self.label_529,4:self.label_540,5:self.label_546,6:self.label_551,7:self.label_770,8:self.label_773,9:self.label_786,10:self.label_787,11:self.label_799,12:self.label_807}
        self.label_barcode_page1 = {0: self.label_124, 1: self.label_516, 2: self.label_523, 3: self.label_530, 4: self.label_536, 5: self.label_548, 6: self.label_552, 7: self.label_769, 8: self.label_775, 9: self.label_780, 10: self.label_790, 11: self.label_797, 12: self.label_801}
        self.label_jiechushijian_page1 = {0: self.label_128, 1: self.label_520, 2: self.label_527, 3: self.label_532, 4: self.label_539, 5: self.label_543, 6: self.label_554, 7: self.label_768, 8: self.label_776, 9: self.label_783, 10: self.label_789, 11: self.label_794, 12: self.label_803}
        self.label_jieyongren_page1 = {0: self.label_132, 1: self.label_517, 2: self.label_524, 3: self.label_528, 4: self.label_537, 5: self.label_542, 6: self.label_555, 7: self.label_771, 8: self.label_777, 9: self.label_785, 10: self.label_792, 11: self.label_795, 12: self.label_805}
        self.label_jieyongren_gonghao_page1 = {0: self.label_133, 1: self.label_515, 2: self.label_522, 3: self.label_533, 4: self.label_538, 5: self.label_545, 6: self.label_556, 7: self.label_767, 8: self.label_778, 9: self.label_784, 10: self.label_793, 11: self.label_798, 12: self.label_802}
        self.label_projectname_page2 = {0: self.label_101, 1: self.label_102, 2: self.label_103, 3: self.label_104, 4: self.label_118, 5: self.label_210, 6: self.label_212, 7: self.label_235, 8: self.label_242, 9: self.label_245, 10: self.label_254, 11: self.label_251, 12: self.label_264}
        self.label_partno_page2={0:self.label_139,1:self.label_151,2:self.label_161,3:self.label_180,4:self.label_183,5:self.label_202,6:self.label_230,7:self.label_237,8:self.label_243,9:self.label_246,10:self.label_256,11:self.label_261,12:self.label_265}
        self.label_partname_page2={0:self.label_146,1:self.label_152,2:self.label_163,3:self.label_182,4:self.label_190,5:self.label_205,6:self.label_231,7:self.label_236,8:self.label_240,9:self.label_249,10:self.label_259,11:self.label_262,12:self.label_266}
        self.label_barcode_page2 = {0: self.label_142, 1: self.label_147, 2: self.label_159, 3: self.label_172, 4: self.label_189, 5: self.label_198, 6: self.label_232, 7: self.label_234, 8: self.label_241, 9: self.label_250, 10: self.label_253, 11: self.label_252, 12: self.label_263}
        self.label_jierushijian_page2 = {0: self.label_140, 1: self.label_158, 2: self.label_169, 3: self.label_176, 4: self.label_195, 5: self.label_201, 6: self.label_211, 7: self.label_238, 8: self.label_244, 9: self.label_248, 10: self.label_258, 11: self.label_260, 12: self.label_268}
        self.label_guazhangren_page2 = {0: self.label_135, 1: self.label_154, 2: self.label_167, 3: self.label_174, 4: self.label_193, 5: self.label_206, 6: self.label_229, 7: self.label_233, 8: self.label_239, 9: self.label_247, 10: self.label_257, 11: self.label_255, 12: self.label_267}
        self.label_projectname_page3 = {0: self.label_114, 1: self.label_115, 2: self.label_117, 3: self.label_116, 4: self.label_284, 5: self.label_287, 6: self.label_295, 7: self.label_297, 8: self.label_305, 9: self.label_307, 10: self.label_315, 11: self.label_317, 12: self.label_650}
        self.label_partname_page3={0:self.label_222,1:self.label_272,2:self.label_278,3:self.label_279,4:self.label_282,5:self.label_290,6:self.label_294,7:self.label_300,8:self.label_304,9:self.label_310,10:self.label_314,11:self.label_320,12:self.label_649}
        self.label_barcode_page3 = {0: self.label_224, 1: self.label_270, 2: self.label_274, 3: self.label_280, 4: self.label_283, 5: self.label_291, 6: self.label_292, 7: self.label_301, 8: self.label_302, 9: self.label_311, 10: self.label_312, 11: self.label_321, 12: self.label_647}
        self.label_guazhangren_page3 = {0: self.label_269, 1: self.label_271, 2: self.label_277, 3: self.label_276, 4: self.label_285, 5: self.label_288, 6: self.label_293, 7: self.label_299, 8: self.label_303, 9: self.label_309, 10: self.label_313, 11: self.label_319, 12: self.label_648}
        self.label_jieyongren_page3 = {0: self.label_221, 1: self.label_273, 2: self.label_281, 3: self.label_275, 4: self.label_286, 5: self.label_289, 6: self.label_296, 7: self.label_298, 8: self.label_306, 9: self.label_308, 10: self.label_316, 11: self.label_318, 12: self.label_651}
        self.pushButton_7.clicked.connect(self.page1)
        self.pushButton_4.clicked.connect(self.page2)
        self.pushButton_39.clicked.connect(self.chaxun_page2)
        self.pushButton_32.clicked.connect(self.chongzhi_page2)
        self.pushButton_213.clicked.connect(self.xiayiye_page2)
        self.pushButton_212.clicked.connect(self.shangyiye_page2)
        self.pushButton_5.clicked.connect(self.page3)
        self.pushButton_41.clicked.connect(self.chaxun_page3)
        self.pushButton_42.clicked.connect(self.chongzhi_page3)
        self.pushButton_217.clicked.connect(self.xiayiye_page3)
        self.pushButton_216.clicked.connect(self.shangyiye_page3)
        self.pushButton_6.clicked.connect(self.page4)
        self.pushButton_89.clicked.connect(self.page5)
        self.pushButton_90.clicked.connect(self.xuanzhe_ruku_file)
        self.pushButton_91.clicked.connect(self.ruku)
        self.pushButton_92.clicked.connect(self.xuanzhe_chuku_file)
        self.pushButton_93.clicked.connect(self.tuiku)
        self.pushButton_87.clicked.connect(self.xiayiye_page4)
        self.pushButton_88.clicked.connect(self.shangyiye_page4)
        self.pushButton_43.clicked.connect(self.chaxun_page4)
        self.pushButton_44.clicked.connect(self.chongzhi_page4)
        #下面是page4的13个借出button和点击借出时候弹出的界面里的确认button28和取消button34
        self.pushButton_34.clicked.connect(self.hide_frame3)
        self.pushButton_24.clicked.connect(self.show_frame3_page4_button24)
        self.pushButton_47.clicked.connect(self.show_frame3_page4_button47)
        self.pushButton_48.clicked.connect(self.show_frame3_page4_button48)
        self.pushButton_49.clicked.connect(self.show_frame3_page4_button49)
        self.pushButton_50.clicked.connect(self.show_frame3_page4_button50)
        self.pushButton_51.clicked.connect(self.show_frame3_page4_button51)
        self.pushButton_52.clicked.connect(self.show_frame3_page4_button52)
        self.pushButton_26.clicked.connect(self.show_frame3_page4_button26)
        self.pushButton_27.clicked.connect(self.show_frame3_page4_button27)
        self.pushButton_29.clicked.connect(self.show_frame3_page4_button29)
        self.pushButton_30.clicked.connect(self.show_frame3_page4_button30)
        self.pushButton_31.clicked.connect(self.show_frame3_page4_button31)
        self.pushButton_33.clicked.connect(self.show_frame3_page4_button33)
        self.pushButton_28.clicked.connect(self.page4_jiqi_jiechu_queren)
        #上面是page4的13个借出button和点击借出时候弹出的界面里的确认和取消button34
        self.pushButton_12.clicked.connect(self.chongzhi_page1)
        self.pushButton_11.clicked.connect(self.chaxun_page1)
        self.pushButton_204.clicked.connect(self.xiayiye_page1)
        self.pushButton_205.clicked.connect(self.shangyiye_page1)
        # 下面是page1的13个归还button和点击归还时候弹出的界面里的确认button81和取消button82
        self.pushButton_13.clicked.connect(self.show_frame5_page1_button13)
        self.pushButton_14.clicked.connect(self.show_frame5_page1_button14)
        self.pushButton_15.clicked.connect(self.show_frame5_page1_button15)
        self.pushButton_16.clicked.connect(self.show_frame5_page1_button16)
        self.pushButton_19.clicked.connect(self.show_frame5_page1_button19)
        self.pushButton_18.clicked.connect(self.show_frame5_page1_button18)
        self.pushButton_17.clicked.connect(self.show_frame5_page1_button17)
        self.pushButton_20.clicked.connect(self.show_frame5_page1_button20)
        self.pushButton_21.clicked.connect(self.show_frame5_page1_button21)
        self.pushButton_200.clicked.connect(self.show_frame5_page1_button200)
        self.pushButton_201.clicked.connect(self.show_frame5_page1_button201)
        self.pushButton_202.clicked.connect(self.show_frame5_page1_button202)
        self.pushButton_203.clicked.connect(self.show_frame5_page1_button203)
        self.pushButton_82.clicked.connect(self.hide_frame5)
        self.pushButton_81.clicked.connect(self.page1_jiqi_guihuan_queren)
        # 上面是page1的13个归还button和点击归还时候弹出的界面里的确认button81和取消button82

 # ---------------这一部分是UI逻辑分离固定格式---------------------

# ---------------这一部分是设置自定义窗口后，app不能拖动，重启可拖动的方式---------------------
    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.is_moving = True
            self.dragPos = event.globalPos() - self.pos()
            event.accept()
    def mouseMoveEvent(self, event):
        if self.is_moving:
            self.move(event.globalPos() - self.dragPos)
            event.accept()

    # 下面是page1里包含的函数
    def page1(self):
        self.frame_5.hide()
        self.lineEdit.clear()  # 清除条件查找框里的内容
        self.lineEdit_2.clear()  # 清除条件查找框里的内容
        for x in range(0, 13):
            self.label_projectname_page1[x].setText('')
            self.label_partno_page1[x].setText('')
            self.label_partname_page1[x].setText('')
            self.label_barcode_page1[x].setText('')
            self.guihuan_button_page1[x].setText('')
            self.label_jiechushijian_page1[x].setText('')
            self.label_jieyongren_page1[x].setText('')
            self.label_jieyongren_gonghao_page1[x].setText('')
        self.label_37.setText('')#暂无记录那个label
        self.label_37.repaint()
        self.label_811.setText('1')
        self.label_811.repaint()
        self.stackedWidget.setCurrentIndex(0)
        try:
            db = pymysql.connect(host="localhost", user='root', password='jiang111', port=3306, database='lily', connect_timeout=3)
            cusor = db.cursor()
            jiechuqingdan_table = 'jiechuqingdan_' + user
            sql = f"select * from {jiechuqingdan_table}"
            cusor.execute(sql)
            self.jiechu_result_page1 = cusor.fetchall()  # 返回这种(('LCH-2', '4X90S91831', 'CABLE USB-C TO ETHERNET ADP SC10P27625', '2022012111770'), ('LCH-1', '20Y6Z9EBCN', 'THINKPAD X1 EXTREME GEN 4 SVT19 PRC', '2022012621453'))
            self.yuanshi_shuju_page1 = self.jiechu_result_page1  # 这里是为了后面每一次条件查询前先恢复到原始data
            self.zongyeshu_page1, self.yushu_page1 = len(self.jiechu_result_page1) // 13, len(self.jiechu_result_page1) % 13  # //取整，%取余数, 后面会用
            if self.zongyeshu_page1 != 0 and self.yushu_page1 != 0:
                self.zongyeshu_page1 = self.zongyeshu_page1 + 1
            if self.zongyeshu_page1 == 0:
                self.label_813.setText('/1')
                self.label_813.repaint()
            else:
                self.label_813.setText(f'/{self.zongyeshu_page1}')
                self.label_813.repaint()
            if len(self.jiechu_result_page1) == 0:
                self.label_37.setText('暂无记录')
                self.label_37.repaint()
                self.label_119.setText('0')
                self.label_119.repaint()
            else:
                self.label_119.setText(str(len(self.jiechu_result_page1)))
                self.label_119.repaint()
                if len(self.jiechu_result_page1) <= 13:
                    for i in range(0, len(self.jiechu_result_page1)):
                        self.label_projectname_page1[i].setText('   ' + self.jiechu_result_page1[i][0])
                        self.label_partno_page1[i].setText('  ' + self.jiechu_result_page1[i][1])
                        self.label_partname_page1[i].setText('  ' + self.jiechu_result_page1[i][2])
                        self.label_barcode_page1[i].setText(' '+self.jiechu_result_page1[i][3])
                        self.label_jiechushijian_page1[i].setText('    ' + self.jiechu_result_page1[i][4])
                        self.label_jieyongren_page1[i].setText('   ' + self.jiechu_result_page1[i][5])
                        self.label_jieyongren_gonghao_page1[i].setText(' ' + self.jiechu_result_page1[i][6])
                        self.guihuan_button_page1[i].setText('                '+ '归还')
                else:  # 大于13
                    for i in range(0, 13):
                        self.label_projectname_page1[i].setText('   ' + self.jiechu_result_page1[i][0])
                        self.label_partno_page1[i].setText('  ' + self.jiechu_result_page1[i][1])
                        self.label_partname_page1[i].setText('  ' + self.jiechu_result_page1[i][2])
                        self.label_barcode_page1[i].setText(' '+self.jiechu_result_page1[i][3])
                        self.label_jiechushijian_page1[i].setText('    '+self.jiechu_result_page1[i][4])
                        self.label_jieyongren_page1[i].setText('   '+self.jiechu_result_page1[i][5])
                        self.label_jieyongren_gonghao_page1[i].setText(' '+self.jiechu_result_page1[i][6])
                        self.guihuan_button_page1[i].setText('                '+'归还')
                db.close()
        except:
            self.label_37.setText('暂无记录')
            self.label_37.repaint()
            self.label_119.setText('0')#0条记录
            self.label_119.repaint()
        list1_page1 = [self.label_100.text(), self.label_125.text(), self.label_126.text(), self.label_124.text(), self.label_128.text(),self.label_132.text(),self.label_133.text()]  # 第一行对应的label的各项内容project name 等
        list2_page1 = [self.label_518.text(), self.label_514.text(), self.label_519.text(), self.label_516.text(), self.label_520.text(),self.label_517.text(),self.label_515.text()]
        list3_page1 = [self.label_525.text(), self.label_521.text(), self.label_526.text(), self.label_523.text(), self.label_527.text(),self.label_524.text(),self.label_522.text()]
        list4_page1 = [self.label_531.text(), self.label_534.text(), self.label_529.text(), self.label_530.text(), self.label_532.text(), self.label_528.text(), self.label_533.text()]
        list5_page1 = [self.label_535.text(), self.label_541.text(), self.label_540.text(), self.label_536.text(), self.label_539.text(), self.label_537.text(), self.label_538.text()]
        list6_page1 = [self.label_544.text(), self.label_547.text(), self.label_546.text(), self.label_548.text(), self.label_543.text(), self.label_542.text(), self.label_545.text()]
        list7_page1 = [self.label_549.text(), self.label_550.text(), self.label_551.text(), self.label_552.text(), self.label_554.text(), self.label_555.text(), self.label_556.text()]
        list8_page1 = [self.label_772.text(), self.label_766.text(), self.label_770.text(), self.label_769.text(), self.label_768.text(), self.label_771.text(), self.label_767.text()]
        list9_page1 = [self.label_779.text(), self.label_774.text(), self.label_773.text(), self.label_775.text(), self.label_776.text(), self.label_777.text(), self.label_778.text()]
        list10_page1 = [self.label_782.text(), self.label_781.text(), self.label_786.text(), self.label_780.text(), self.label_783.text(), self.label_785.text(), self.label_784.text()]
        list11_page1 = [self.label_788.text(), self.label_791.text(), self.label_787.text(), self.label_790.text(), self.label_789.text(), self.label_792.text(), self.label_793.text()]
        list12_page1 = [self.label_796.text(), self.label_800.text(), self.label_799.text(), self.label_797.text(), self.label_794.text(), self.label_795.text(), self.label_798.text()]
        list13_page1 = [self.label_804.text(), self.label_806.text(), self.label_807.text(), self.label_801.text(), self.label_803.text(), self.label_805.text(), self.label_802.text()]
        self.guihuan_button_duiying_label_page1 = {'pushButton_13': list1_page1, 'pushButton_14': list2_page1, 'pushButton_15': list3_page1, 'pushButton_16': list4_page1, 'pushButton_19': list5_page1, 'pushButton_18': list6_page1, 'pushButton_17': list7_page1, 'pushButton_20': list8_page1, 'pushButton_21': list9_page1, 'pushButton_200': list10_page1, 'pushButton_201': list11_page1, 'pushButton_202': list12_page1, 'pushButton_203': list13_page1}
    def chongzhi_page1(self):
        self.lineEdit.clear()
        self.lineEdit_2.clear()
    def xiayiye_page1(self):#只有大于13的时候，下一页button才有效，这里不用数据库，展示page4(self)的结果
        if int(self.label_811.text()) >=self.zongyeshu_page1:#如果当前页码大于等于总页码，点击不执行任何东西
            pass
        else:
            for x in range(0, 13):#点击下页的时候首先清掉当前页的所有内容，再展示下一页的.
                self.label_projectname_page1[x].setText('')
                self.label_partno_page1[x].setText('')
                self.label_partname_page1[x].setText('')
                self.label_barcode_page1[x].setText('')
                self.guihuan_button_page1[x].setText('')
                self.label_jiechushijian_page1[x].setText('')
                self.label_jieyongren_page1[x].setText('')
                self.label_jieyongren_gonghao_page1[x].setText('')
            yema=int(self.label_811.text())
            self.label_811.setText(str(yema+1))
            self.label_811.repaint()
            current_yema = int(self.label_811.text())
            for y in range(0,13):
                try:
                    self.label_projectname_page1[y].setText('   ' + self.jiechu_result_page1[(current_yema-1)*13+y][0])#y+13代表显示列表的13项开始，依次类推
                    self.label_projectname_page1[y].repaint()
                    self.label_partno_page1[y].setText('  ' + self.jiechu_result_page1[(current_yema-1)*13+y][1])
                    self.label_partno_page1[y].repaint()
                    self.label_partname_page1[y].setText('  ' + self.jiechu_result_page1[(current_yema-1)*13+y][2])
                    self.label_partname_page1[y].repaint()
                    self.label_barcode_page1[y].setText(' ' + self.jiechu_result_page1[(current_yema-1)*13+y][3])
                    self.label_barcode_page1[y].repaint()
                    self.label_jiechushijian_page1[y].setText('    ' + self.jiechu_result_page1[(current_yema-1)*13+y][4])
                    self.label_jiechushijian_page1[y].repaint()
                    self.label_jieyongren_page1[y].setText('   ' + self.jiechu_result_page1[(current_yema-1)*13+y][5])
                    self.label_jieyongren_page1[y].repaint()
                    self.label_jieyongren_gonghao_page1[y].setText(' ' + self.jiechu_result_page1[(current_yema-1)*13+y][6])
                    self.label_jieyongren_gonghao_page1[y].repaint()
                    self.guihuan_button_page1[y].setText('                ' + '归还')
                except:###这里的作用是如果最后一页没有13个选项内容，前面肯定就会报错，这里就跳过报错，最后一页有多少就显示多少
                    pass
        list1_page1 = [self.label_100.text(), self.label_125.text(), self.label_126.text(), self.label_124.text(), self.label_128.text(),self.label_132.text(),self.label_133.text()]  # 第一行对应的label的各项内容project name 等
        list2_page1 = [self.label_518.text(), self.label_514.text(), self.label_519.text(), self.label_516.text(), self.label_520.text(),self.label_517.text(),self.label_515.text()]
        list3_page1 = [self.label_525.text(), self.label_521.text(), self.label_526.text(), self.label_523.text(), self.label_527.text(),self.label_524.text(),self.label_522.text()]
        list4_page1 = [self.label_531.text(), self.label_534.text(), self.label_529.text(), self.label_530.text(), self.label_532.text(), self.label_528.text(), self.label_533.text()]
        list5_page1 = [self.label_535.text(), self.label_541.text(), self.label_540.text(), self.label_536.text(), self.label_539.text(), self.label_537.text(), self.label_538.text()]
        list6_page1 = [self.label_544.text(), self.label_547.text(), self.label_546.text(), self.label_548.text(), self.label_543.text(), self.label_542.text(), self.label_545.text()]
        list7_page1 = [self.label_549.text(), self.label_550.text(), self.label_551.text(), self.label_552.text(), self.label_554.text(), self.label_555.text(), self.label_556.text()]
        list8_page1 = [self.label_772.text(), self.label_766.text(), self.label_770.text(), self.label_769.text(), self.label_768.text(), self.label_771.text(), self.label_767.text()]
        list9_page1 = [self.label_779.text(), self.label_774.text(), self.label_773.text(), self.label_775.text(), self.label_776.text(), self.label_777.text(), self.label_778.text()]
        list10_page1 = [self.label_782.text(), self.label_781.text(), self.label_786.text(), self.label_780.text(), self.label_783.text(), self.label_785.text(), self.label_784.text()]
        list11_page1 = [self.label_788.text(), self.label_791.text(), self.label_787.text(), self.label_790.text(), self.label_789.text(), self.label_792.text(), self.label_793.text()]
        list12_page1 = [self.label_796.text(), self.label_800.text(), self.label_799.text(), self.label_797.text(), self.label_794.text(), self.label_795.text(), self.label_798.text()]
        list13_page1 = [self.label_804.text(), self.label_806.text(), self.label_807.text(), self.label_801.text(), self.label_803.text(), self.label_805.text(), self.label_802.text()]
        self.guihuan_button_duiying_label_page1 = {'pushButton_13': list1_page1, 'pushButton_14': list2_page1, 'pushButton_15': list3_page1, 'pushButton_16': list4_page1, 'pushButton_19': list5_page1, 'pushButton_18': list6_page1, 'pushButton_17': list7_page1, 'pushButton_20': list8_page1, 'pushButton_21': list9_page1, 'pushButton_200': list10_page1, 'pushButton_201': list11_page1, 'pushButton_202': list12_page1, 'pushButton_203': list13_page1}
    def shangyiye_page1(self):  # 只有大于13的时候，下一页button才有效，这里不用数据库，展示page4(self)的结果
        if int(self.label_811.text()) ==1:  # 如果当前页码等于1的时候点击不执行任何动作
            pass
        else:
            for x in range(0, 13):  # 点击下页的时候首先清掉当前页的所有内容，再展示下一页的.
                self.label_projectname_page1[x].setText('')
                self.label_partno_page1[x].setText('')
                self.label_partname_page1[x].setText('')
                self.label_barcode_page1[x].setText('')
                self.guihuan_button_page1[x].setText('')
                self.label_jiechushijian_page1[x].setText('')
                self.label_jieyongren_page1[x].setText('')
                self.label_jieyongren_gonghao_page1[x].setText('')
            yema = int(self.label_811.text())
            self.label_811.setText(str(yema - 1))
            self.label_811.repaint()
            current_yema = int(self.label_811.text())
            for y in range(0, 13):#上一页肯定会显示满，所以不需要用try
                self.label_projectname_page1[y].setText('   ' + self.jiechu_result_page1[(current_yema - 1) * 13 + y][0])  # y+13代表显示列表的13项开始，依次类推
                self.label_projectname_page1[y].repaint()
                self.label_partno_page1[y].setText('  ' + self.jiechu_result_page1[(current_yema - 1) * 13 + y][1])
                self.label_partno_page1[y].repaint()
                self.label_partname_page1[y].setText('  ' + self.jiechu_result_page1[(current_yema - 1) * 13 + y][2])
                self.label_partname_page1[y].repaint()
                self.label_barcode_page1[y].setText(' ' + self.jiechu_result_page1[(current_yema - 1) * 13 + y][3])
                self.label_barcode_page1[y].repaint()
                self.label_jiechushijian_page1[y].setText('    ' + self.jiechu_result_page1[(current_yema - 1) * 13 + y][4])
                self.label_jiechushijian_page1[y].repaint()
                self.label_jieyongren_page1[y].setText('   ' + self.jiechu_result_page1[(current_yema - 1) * 13 + y][5])
                self.label_jieyongren_page1[y].repaint()
                self.label_jieyongren_gonghao_page1[y].setText(' ' + self.jiechu_result_page1[(current_yema - 1) * 13 + y][6])
                self.label_jieyongren_gonghao_page1[y].repaint()
                self.guihuan_button_page1[y].setText('                ' + '归还')
        list1_page1 = [self.label_100.text(), self.label_125.text(), self.label_126.text(), self.label_124.text(), self.label_128.text(),self.label_132.text(),self.label_133.text()]  # 第一行对应的label的各项内容project name 等
        list2_page1 = [self.label_518.text(), self.label_514.text(), self.label_519.text(), self.label_516.text(), self.label_520.text(),self.label_517.text(),self.label_515.text()]
        list3_page1 = [self.label_525.text(), self.label_521.text(), self.label_526.text(), self.label_523.text(), self.label_527.text(),self.label_524.text(),self.label_522.text()]
        list4_page1 = [self.label_531.text(), self.label_534.text(), self.label_529.text(), self.label_530.text(), self.label_532.text(), self.label_528.text(), self.label_533.text()]
        list5_page1 = [self.label_535.text(), self.label_541.text(), self.label_540.text(), self.label_536.text(), self.label_539.text(), self.label_537.text(), self.label_538.text()]
        list6_page1 = [self.label_544.text(), self.label_547.text(), self.label_546.text(), self.label_548.text(), self.label_543.text(), self.label_542.text(), self.label_545.text()]
        list7_page1 = [self.label_549.text(), self.label_550.text(), self.label_551.text(), self.label_552.text(), self.label_554.text(), self.label_555.text(), self.label_556.text()]
        list8_page1 = [self.label_772.text(), self.label_766.text(), self.label_770.text(), self.label_769.text(), self.label_768.text(), self.label_771.text(), self.label_767.text()]
        list9_page1 = [self.label_779.text(), self.label_774.text(), self.label_773.text(), self.label_775.text(), self.label_776.text(), self.label_777.text(), self.label_778.text()]
        list10_page1 = [self.label_782.text(), self.label_781.text(), self.label_786.text(), self.label_780.text(), self.label_783.text(), self.label_785.text(), self.label_784.text()]
        list11_page1 = [self.label_788.text(), self.label_791.text(), self.label_787.text(), self.label_790.text(), self.label_789.text(), self.label_792.text(), self.label_793.text()]
        list12_page1 = [self.label_796.text(), self.label_800.text(), self.label_799.text(), self.label_797.text(), self.label_794.text(), self.label_795.text(), self.label_798.text()]
        list13_page1 = [self.label_804.text(), self.label_806.text(), self.label_807.text(), self.label_801.text(), self.label_803.text(), self.label_805.text(), self.label_802.text()]
        self.guihuan_button_duiying_label_page1 = {'pushButton_13': list1_page1, 'pushButton_14': list2_page1, 'pushButton_15': list3_page1, 'pushButton_16': list4_page1, 'pushButton_19': list5_page1, 'pushButton_18': list6_page1, 'pushButton_17': list7_page1, 'pushButton_20': list8_page1, 'pushButton_21': list9_page1, 'pushButton_200': list10_page1, 'pushButton_201': list11_page1, 'pushButton_202': list12_page1, 'pushButton_203': list13_page1}
    def chaxun_page1(self):#查询只设置3种情况，要么2个标签都有内容，要么其中一个有，否则查询函数为pass
        self.jiechu_result_page1=self.yuanshi_shuju_page1#每次复原后再查询
        if self.lineEdit.text()!='' and self.lineEdit_2.text()=='':#project name不为空
            page1_chaxun_project_list=[]#准备一个新的列表，筛选条件后，插入新的元素，依赖前面生成的所有的总的返回结果self.guazhang_result_page4，因为元组不能插入元素，写写列表后面转元组
            for x in range(0, 13):
                self.label_projectname_page1[x].setText('')
                self.label_partno_page1[x].setText('')
                self.label_partname_page1[x].setText('')
                self.label_barcode_page1[x].setText('')
                self.guihuan_button_page1[x].setText('')
                self.label_jiechushijian_page1[x].setText('')
                self.label_jieyongren_page1[x].setText('')
                self.label_jieyongren_gonghao_page1[x].setText('')
            self.label_37.setText('')#88是显示暂无记录那个标签
            self.label_37.repaint()
            self.label_811.setText('1')
            self.label_811.repaint()
            #self.guazhang=(('LCH-2', '4X90S91831', 'CABLE USB-C TO ETHERNET ADP SC10P27625', '2022012111770'),('LCH-1', '20Y6Z9EBCN', 'THINKPAD X1 EXTREME GEN 4 SVT19 PRC', '2022012621453'))
            for each in self.jiechu_result_page1:
                if self.lineEdit.text() in each[0]:#如果搜索的名字被包含在project name里
                    page1_chaxun_project_list.append(each)
            self.jiechu_result_page1=tuple(page1_chaxun_project_list)
            self.zongyeshu_page1, self.yushu_page1 = len(self.jiechu_result_page1) // 13, len(self.jiechu_result_page1) % 13  # //取整，%取余数, 后面会用
            if self.zongyeshu_page1 != 0 and self.yushu_page1 != 0:
                self.zongyeshu_page1 = self.zongyeshu_page1 + 1
            if self.zongyeshu_page1 == 0:
                self.label_813.setText('/1')
                self.label_813.repaint()
            else:
                self.label_813.setText(f'/{self.zongyeshu_page1}')
                self.label_813.repaint()
            if len(self.jiechu_result_page1) == 0:
                self.label_37.setText('暂无记录')
                self.label_37.repaint()
                self.label_119.setText('0')
                self.label_119.repaint()
            else:
                self.label_119.setText(str(len(self.jiechu_result_page1)))
                self.label_119.repaint()
                if len(self.jiechu_result_page1) <= 13:
                    for i in range(0, len(self.jiechu_result_page1)):
                        self.label_projectname_page1[i].setText('   ' + self.jiechu_result_page1[i][0])
                        self.label_partno_page1[i].setText('  ' + self.jiechu_result_page1[i][1])
                        self.label_partname_page1[i].setText('  ' + self.jiechu_result_page1[i][2])
                        self.label_barcode_page1[i].setText(' '+self.jiechu_result_page1[i][3])
                        self.label_jiechushijian_page1[i].setText('    ' + self.jiechu_result_page1[i][4])
                        self.label_jieyongren_page1[i].setText('   ' + self.jiechu_result_page1[i][5])
                        self.label_jieyongren_gonghao_page1[i].setText(' ' + self.jiechu_result_page1[i][6])
                        self.guihuan_button_page1[i].setText('                '+ '归还')
                else:  # 大于13
                    for i in range(0, 13):
                        self.label_projectname_page1[i].setText('   ' + self.jiechu_result_page1[i][0])
                        self.label_partno_page1[i].setText('  ' + self.jiechu_result_page1[i][1])
                        self.label_partname_page1[i].setText('  ' + self.jiechu_result_page1[i][2])
                        self.label_barcode_page1[i].setText(' '+self.jiechu_result_page1[i][3])
                        self.label_jiechushijian_page1[i].setText('    ' + self.jiechu_result_page1[i][4])
                        self.label_jieyongren_page1[i].setText('   ' + self.jiechu_result_page1[i][5])
                        self.label_jieyongren_gonghao_page1[i].setText(' ' + self.jiechu_result_page1[i][6])
                        self.guihuan_button_page1[i].setText('                '+ '归还')
        elif self.lineEdit.text() == '' and self.lineEdit_2.text() != '':  # part name不为空
            page1_chaxun_partname_list = []  # 准备一个新的列表，筛选条件后，插入新的元素，依赖前面生成的所有的总的返回结果self.guazhang_result_page4，因为元组不能插入元素，写写列表后面转元组
            for x in range(0, 13):
                self.label_projectname_page1[x].setText('')
                self.label_partno_page1[x].setText('')
                self.label_partname_page1[x].setText('')
                self.label_barcode_page1[x].setText('')
                self.guihuan_button_page1[x].setText('')
                self.label_jiechushijian_page1[x].setText('')
                self.label_jieyongren_page1[x].setText('')
                self.label_jieyongren_gonghao_page1[x].setText('')
            self.label_37.setText('')  # 88是显示暂无记录那个标签
            self.label_37.repaint()
            self.label_811.setText('1')
            self.label_811.repaint()
            # self.guazhang=(('LCH-2', '4X90S91831', 'CABLE USB-C TO ETHERNET ADP SC10P27625', '2022012111770'),('LCH-1', '20Y6Z9EBCN', 'THINKPAD X1 EXTREME GEN 4 SVT19 PRC', '2022012621453'))
            for each in self.jiechu_result_page1:
                if self.lineEdit_2.text() in each[2]:  # 如果搜索的名字被包含在PARTname里
                    page1_chaxun_partname_list.append(each)
            self.jiechu_result_page1 = tuple(page1_chaxun_partname_list)
            self.zongyeshu_page1, self.yushu_page1 = len(self.jiechu_result_page1) // 13, len(self.jiechu_result_page1) % 13  # //取整，%取余数, 后面会用
            if self.zongyeshu_page1 != 0 and self.yushu_page1 != 0:
                self.zongyeshu_page1 = self.zongyeshu_page1 + 1
            if self.zongyeshu_page1 == 0:
                self.label_813.setText('/1')
                self.label_813.repaint()
            else:
                self.label_813.setText(f'/{self.zongyeshu_page1}')
                self.label_813.repaint()
            if len(self.jiechu_result_page1) == 0:
                self.label_37.setText('暂无记录')
                self.label_37.repaint()
                self.label_119.setText('0')
                self.label_119.repaint()
            else:
                self.label_119.setText(str(len(self.jiechu_result_page1)))
                self.label_119.repaint()
                if len(self.jiechu_result_page1) <= 13:
                    for i in range(0, len(self.jiechu_result_page1)):
                        self.label_projectname_page1[i].setText('   ' + self.jiechu_result_page1[i][0])
                        self.label_partno_page1[i].setText('  ' + self.jiechu_result_page1[i][1])
                        self.label_partname_page1[i].setText('  ' + self.jiechu_result_page1[i][2])
                        self.label_barcode_page1[i].setText(' '+self.jiechu_result_page1[i][3])
                        self.label_jiechushijian_page1[i].setText('    ' + self.jiechu_result_page1[i][4])
                        self.label_jieyongren_page1[i].setText('   ' + self.jiechu_result_page1[i][5])
                        self.label_jieyongren_gonghao_page1[i].setText(' ' + self.jiechu_result_page1[i][6])
                        self.guihuan_button_page1[i].setText('                '+ '归还')
                else:  # 大于13
                    for i in range(0, 13):
                        self.label_projectname_page1[i].setText('   ' + self.jiechu_result_page1[i][0])
                        self.label_partno_page1[i].setText('  ' + self.jiechu_result_page1[i][1])
                        self.label_partname_page1[i].setText('  ' + self.jiechu_result_page1[i][2])
                        self.label_barcode_page1[i].setText(' '+self.jiechu_result_page1[i][3])
                        self.label_jiechushijian_page1[i].setText('    ' + self.jiechu_result_page1[i][4])
                        self.label_jieyongren_page1[i].setText('   ' + self.jiechu_result_page1[i][5])
                        self.label_jieyongren_gonghao_page1[i].setText(' ' + self.jiechu_result_page1[i][6])
                        self.guihuan_button_page1[i].setText('                '+ '归还')
        elif self.lineEdit.text() != '' and self.lineEdit_2.text() != '':  # part 和 project name不为空
            page1_chaxun_part_project_name_list = []  # 准备一个新的列表，筛选条件后，插入新的元素，依赖前面生成的所有的总的返回结果self.guazhang_result_page4，因为元组不能插入元素，写写列表后面转元组
            for x in range(0, 13):
                self.label_projectname_page1[x].setText('')
                self.label_partno_page1[x].setText('')
                self.label_partname_page1[x].setText('')
                self.label_barcode_page1[x].setText('')
                self.guihuan_button_page1[x].setText('')
                self.label_jiechushijian_page1[x].setText('')
                self.label_jieyongren_page1[x].setText('')
                self.label_jieyongren_gonghao_page1[x].setText('')
            self.label_37.setText('')  # 88是显示暂无记录那个标签
            self.label_37.repaint()
            self.label_811.setText('1')
            self.label_811.repaint()
            # self.guazhang=(('LCH-2', '4X90S91831', 'CABLE USB-C TO ETHERNET ADP SC10P27625', '2022012111770'),('LCH-1', '20Y6Z9EBCN', 'THINKPAD X1 EXTREME GEN 4 SVT19 PRC', '2022012621453'))
            for each in self.jiechu_result_page1:
                if self.lineEdit.text() in each[2] and self.lineEdit_2.text() in each[0]:  # 如果搜索的名字被包含在PARTname里
                    page1_chaxun_part_project_name_list.append(each)
            self.jiechu_result_page1 = tuple(page1_chaxun_part_project_name_list)
            self.zongyeshu_page1, self.yushu_page1 = len(self.jiechu_result_page1) // 13, len(self.jiechu_result_page1) % 13  # //取整，%取余数, 后面会用
            if self.zongyeshu_page1 != 0 and self.yushu_page1 != 0:
                self.zongyeshu_page1 = self.zongyeshu_page1 + 1
            if self.zongyeshu_page1 == 0:
                self.label_813.setText('/1')
                self.label_813.repaint()
            else:
                self.label_813.setText(f'/{self.zongyeshu_page1}')
                self.label_813.repaint()
            if len(self.jiechu_result_page1) == 0:
                self.label_37.setText('暂无记录')
                self.label_37.repaint()
                self.label_119.setText('0')
                self.label_119.repaint()
            else:
                self.label_119.setText(str(len(self.jiechu_result_page1)))
                self.label_119.repaint()
                if len(self.jiechu_result_page1) <= 13:
                    for i in range(0, len(self.jiechu_result_page1)):
                        self.label_projectname_page1[i].setText('   ' + self.jiechu_result_page1[i][0])
                        self.label_partno_page1[i].setText('  ' + self.jiechu_result_page1[i][1])
                        self.label_partname_page1[i].setText('  ' + self.jiechu_result_page1[i][2])
                        self.label_barcode_page1[i].setText(' '+self.jiechu_result_page1[i][3])
                        self.label_jiechushijian_page1[i].setText('    ' + self.jiechu_result_page1[i][4])
                        self.label_jieyongren_page1[i].setText('   ' + self.jiechu_result_page1[i][5])
                        self.label_jieyongren_gonghao_page1[i].setText(' ' + self.jiechu_result_page1[i][6])
                        self.guihuan_button_page1[i].setText('                '+ '归还')
                else:  # 大于13
                    for i in range(0, 13):
                        self.label_projectname_page1[i].setText('   ' + self.jiechu_result_page1[i][0])
                        self.label_partno_page1[i].setText('  ' + self.jiechu_result_page1[i][1])
                        self.label_partname_page1[i].setText('  ' + self.jiechu_result_page1[i][2])
                        self.label_barcode_page1[i].setText(' '+self.jiechu_result_page1[i][3])
                        self.label_jiechushijian_page1[i].setText('    ' + self.jiechu_result_page1[i][4])
                        self.label_jieyongren_page1[i].setText('   ' + self.jiechu_result_page1[i][5])
                        self.label_jieyongren_gonghao_page1[i].setText(' ' + self.jiechu_result_page1[i][6])
                        self.guihuan_button_page1[i].setText('                '+ '归还')
        else:#如果2个都为空, 就什么都不做
            pass
        list1_page1 = [self.label_100.text(), self.label_125.text(), self.label_126.text(), self.label_124.text(), self.label_128.text(),self.label_132.text(),self.label_133.text()]  # 第一行对应的label的各项内容project name 等
        list2_page1 = [self.label_518.text(), self.label_514.text(), self.label_519.text(), self.label_516.text(), self.label_520.text(),self.label_517.text(),self.label_515.text()]
        list3_page1 = [self.label_525.text(), self.label_521.text(), self.label_526.text(), self.label_523.text(), self.label_527.text(),self.label_524.text(),self.label_522.text()]
        list4_page1 = [self.label_531.text(), self.label_534.text(), self.label_529.text(), self.label_530.text(), self.label_532.text(), self.label_528.text(), self.label_533.text()]
        list5_page1 = [self.label_535.text(), self.label_541.text(), self.label_540.text(), self.label_536.text(), self.label_539.text(), self.label_537.text(), self.label_538.text()]
        list6_page1 = [self.label_544.text(), self.label_547.text(), self.label_546.text(), self.label_548.text(), self.label_543.text(), self.label_542.text(), self.label_545.text()]
        list7_page1 = [self.label_549.text(), self.label_550.text(), self.label_551.text(), self.label_552.text(), self.label_554.text(), self.label_555.text(), self.label_556.text()]
        list8_page1 = [self.label_772.text(), self.label_766.text(), self.label_770.text(), self.label_769.text(), self.label_768.text(), self.label_771.text(), self.label_767.text()]
        list9_page1 = [self.label_779.text(), self.label_774.text(), self.label_773.text(), self.label_775.text(), self.label_776.text(), self.label_777.text(), self.label_778.text()]
        list10_page1 = [self.label_782.text(), self.label_781.text(), self.label_786.text(), self.label_780.text(), self.label_783.text(), self.label_785.text(), self.label_784.text()]
        list11_page1 = [self.label_788.text(), self.label_791.text(), self.label_787.text(), self.label_790.text(), self.label_789.text(), self.label_792.text(), self.label_793.text()]
        list12_page1 = [self.label_796.text(), self.label_800.text(), self.label_799.text(), self.label_797.text(), self.label_794.text(), self.label_795.text(), self.label_798.text()]
        list13_page1 = [self.label_804.text(), self.label_806.text(), self.label_807.text(), self.label_801.text(), self.label_803.text(), self.label_805.text(), self.label_802.text()]
        self.guihuan_button_duiying_label_page1 = {'pushButton_13': list1_page1, 'pushButton_14': list2_page1, 'pushButton_15': list3_page1, 'pushButton_16': list4_page1, 'pushButton_19': list5_page1, 'pushButton_18': list6_page1, 'pushButton_17': list7_page1, 'pushButton_20': list8_page1, 'pushButton_21': list9_page1, 'pushButton_200': list10_page1, 'pushButton_201': list11_page1, 'pushButton_202': list12_page1, 'pushButton_203': list13_page1}
    def show_frame5_page1_button13(self):#点击借出的时候显示借出框
        self.frame_5.show()
        self.label_15.setText('')
        self.label_15.setText('pushButton_13')
        self.label_654.setText('')
    def show_frame5_page1_button14(self):#点击借出的时候显示借出框
        self.frame_5.show()
        self.label_15.setText('')
        self.label_15.setText('pushButton_14')
        self.label_654.setText('')
    def show_frame5_page1_button15(self):  # 点击借出的时候显示借出框
        self.frame_5.show()
        self.label_15.setText('')
        self.label_15.setText('pushButton_15')
        self.label_654.setText('')
    def show_frame5_page1_button16(self):  # 点击借出的时候显示借出框
        self.frame_5.show()
        self.label_15.setText('')
        self.label_15.setText('pushButton_16')
        self.label_654.setText('')
    def show_frame5_page1_button19(self):  # 点击借出的时候显示借出框
        self.frame_5.show()
        self.label_15.setText('')
        self.label_15.setText('pushButton_19')
        self.label_654.setText('')
    def show_frame5_page1_button18(self):  # 点击借出的时候显示借出框
        self.frame_5.show()
        self.label_15.setText('')
        self.label_15.setText('pushButton_18')
        self.label_654.setText('')
    def show_frame5_page1_button17(self):  # 点击借出的时候显示借出框
        self.frame_5.show()
        self.label_15.setText('')
        self.label_15.setText('pushButton_17')
        self.label_654.setText('')
    def show_frame5_page1_button20(self):  # 点击借出的时候显示借出框
        self.frame_5.show()
        self.label_15.setText('')
        self.label_15.setText('pushButton_20')
        self.label_654.setText('')
    def show_frame5_page1_button21(self):  # 点击借出的时候显示借出框
        self.frame_5.show()
        self.label_15.setText('')
        self.label_15.setText('pushButton_21')
        self.label_654.setText('')
    def show_frame5_page1_button200(self):  # 点击借出的时候显示借出框
        self.frame_5.show()
        self.label_15.setText('')
        self.label_15.setText('pushButton_200')
        self.label_654.setText('')
    def show_frame5_page1_button201(self):  # 点击借出的时候显示借出框
        self.frame_5.show()
        self.label_15.setText('')
        self.label_15.setText('pushButton_201')
        self.label_654.setText('')
    def show_frame5_page1_button202(self):  # 点击借出的时候显示借出框
        self.frame_5.show()
        self.label_15.setText('')
        self.label_15.setText('pushButton_202')
        self.label_654.setText('')
    def show_frame5_page1_button203(self):  # 点击借出的时候显示借出框
        self.frame_5.show()
        self.label_15.setText('')
        self.label_15.setText('pushButton_203')
        self.label_654.setText('')
    def hide_frame5(self):
        self.frame_5.hide()
    def page1_jiqi_guihuan_queren(self):# 15——label是记录button编号的label标签，颜色和背景一样看不出来哈哈
        jiechuqingdan_table = 'jiechuqingdan_' + user#数据库表格
        guazhang_table='guazhang_'+user
        self.label_654.setText('')
        try:
            db = pymysql.connect(host="localhost", user='root', password='jiang111', port=3306, database='lily', connect_timeout=4)
            cusor = db.cursor()
            try:
                Barcode=self.guihuan_button_duiying_label_page1[self.label_15.text()][3].replace(' ','')#label的内容有空格
                sql=f"UPDATE {guazhang_table} set Status='未借出' WHERE Barcode='{Barcode}'"
                sql_1 = f"DELETE FROM {jiechuqingdan_table} WHERE Barcode='{Barcode}'"
                jieyongren_gonghao_page1=self.guihuan_button_duiying_label_page1[self.label_15.text()][6].replace(' ','')
                sql_3 = f"UPDATE {self.wuzichazhao} set jieyong_ren='无' WHERE Barcode='{Barcode}'"
                cusor.execute(sql)
                cusor.execute(sql_1)
                cusor.execute(sql_3)
                jieru_qingdan_table_page1='jieruqingdan_'+jieyongren_gonghao_page1
                if jieyongren_gonghao_page1 in gonghao_all:
                    sql_2=f"DELETE FROM {jieru_qingdan_table_page1} WHERE Barcode='{Barcode}'"
                    cusor.execute(sql_2)
                db.commit()
                self.frame_5.hide()
            except:
                db.rollback()
                self.label_654.setText('归还失败')
        except:
                self.label_654.setText('归还失败')
    # 上面是page1里包含的函数

    # 下面是page2里包含的函数
    def page2(self):
        self.lineEdit_3.clear()  # 清除条件查找框里的内容
        self.lineEdit_4.clear()  # 清除条件查找框里的内容
        for x in range(0, 13):
            self.label_projectname_page2[x].setText('')
            self.label_partno_page2[x].setText('')
            self.label_partname_page2[x].setText('')
            self.label_barcode_page2[x].setText('')
            self.label_jierushijian_page2[x].setText('')
            self.label_guazhangren_page2[x].setText('')
        self.label_41.setText('')#暂无记录那个label
        self.label_41.repaint()
        self.label_864.setText('1')
        self.label_864.repaint()
        self.stackedWidget.setCurrentIndex(1)
        try:
            db = pymysql.connect(host="localhost", user='root', password='jiang111', port=3306, database='lily', connect_timeout=3)
            cusor = db.cursor()
            jieruqingdan_table = 'jieruqingdan_' + user
            sql = f"select * from {jieruqingdan_table}"
            cusor.execute(sql)
            self.jieru_result_page2 = cusor.fetchall()  # 返回这种(('LCH-2', '4X90S91831', 'CABLE USB-C TO ETHERNET ADP SC10P27625', '2022012111770'), ('LCH-1', '20Y6Z9EBCN', 'THINKPAD X1 EXTREME GEN 4 SVT19 PRC', '2022012621453'))
            self.yuanshi_shuju_page2 = self.jieru_result_page2  # 这里是为了后面每一次条件查询前先恢复到原始data
            self.zongyeshu_page2, self.yushu_page2 = len(self.jieru_result_page2) // 13, len(self.jieru_result_page2) % 13  # //取整，%取余数, 后面会用
            if self.zongyeshu_page2 != 0 and self.yushu_page2 != 0:
                self.zongyeshu_page2 = self.zongyeshu_page2 + 1
            if self.zongyeshu_page2 == 0:
                self.label_862.setText('/1')
                self.label_862.repaint()
            else:
                self.label_862.setText(f'/{self.zongyeshu_page2}')
                self.label_862.repaint()
            if len(self.jieru_result_page2) == 0:
                self.label_41.setText('暂无记录')
                self.label_41.repaint()
                self.label_471.setText('0')
                self.label_471.repaint()
            else:
                self.label_471.setText(str(len(self.jieru_result_page2)))
                self.label_471.repaint()
                if len(self.jieru_result_page2) <= 13:
                    for i in range(0, len(self.jieru_result_page2)):
                        self.label_projectname_page2[i].setText('   ' + self.jieru_result_page2[i][0])
                        self.label_partno_page2[i].setText('  ' + self.jieru_result_page2[i][1])
                        self.label_partname_page2[i].setText('      ' + self.jieru_result_page2[i][2])
                        self.label_barcode_page2[i].setText('   '+self.jieru_result_page2[i][3])
                        self.label_jierushijian_page2[i].setText(' ' + self.jieru_result_page2[i][4])
                        self.label_guazhangren_page2[i].setText('     ' + self.jieru_result_page2[i][5])
                else:  # 大于13
                    for i in range(0, 13):
                        self.label_projectname_page2[i].setText('   ' + self.jieru_result_page2[i][0])
                        self.label_partno_page2[i].setText('  ' + self.jieru_result_page2[i][1])
                        self.label_partname_page2[i].setText('      ' + self.jieru_result_page2[i][2])
                        self.label_barcode_page2[i].setText('   '+self.jieru_result_page2[i][3])
                        self.label_jierushijian_page2[i].setText(' ' + self.jieru_result_page2[i][4])
                        self.label_guazhangren_page2[i].setText('     ' + self.jieru_result_page2[i][5])
                db.close()
        except:
            self.label_41.setText('暂无记录')
            self.label_41.repaint()
            self.label_471.setText('0')#0条记录
            self.label_471.repaint()
    def chaxun_page2(self):#查询只设置3种情况，要么2个标签都有内容，要么其中一个有，否则查询函数为pass
        self.jieru_result_page2=self.yuanshi_shuju_page2#每次复原后再查询
        if self.lineEdit_4.text()!='' and self.lineEdit_3.text()=='':#project name不为空
            page2_chaxun_project_list=[]#准备一个新的列表，筛选条件后，插入新的元素，依赖前面生成的所有的总的返回结果self.guazhang_result_page4，因为元组不能插入元素，写写列表后面转元组
            for x in range(0, 13):
                self.label_projectname_page2[x].setText('')
                self.label_partno_page2[x].setText('')
                self.label_partname_page2[x].setText('')
                self.label_barcode_page2[x].setText('')
                self.label_jierushijian_page2[x].setText('')
                self.label_guazhangren_page2[x].setText('')
            self.label_41.setText('')#88是显示暂无记录那个标签
            self.label_41.repaint()
            self.label_864.setText('1')
            self.label_864.repaint()
            #self.guazhang=(('LCH-2', '4X90S91831', 'CABLE USB-C TO ETHERNET ADP SC10P27625', '2022012111770'),('LCH-1', '20Y6Z9EBCN', 'THINKPAD X1 EXTREME GEN 4 SVT19 PRC', '2022012621453'))
            for each in self.jieru_result_page2:
                if self.lineEdit_4.text() in each[0]:#如果搜索的名字被包含在project name里
                    page2_chaxun_project_list.append(each)
            self.jieru_result_page2=tuple(page2_chaxun_project_list)
            self.zongyeshu_page2, self.yushu_page2 = len(self.jieru_result_page2) // 13, len(self.jieru_result_page2) % 13  # //取整，%取余数, 后面会用
            if self.zongyeshu_page2 != 0 and self.yushu_page2 != 0:
                self.zongyeshu_page2 = self.zongyeshu_page2 + 1
            if self.zongyeshu_page2 == 0:
                self.label_862.setText('/1')
                self.label_862.repaint()
            else:
                self.label_862.setText(f'/{self.zongyeshu_page2}')
                self.label_862.repaint()
            if len(self.jieru_result_page2) == 0:
                self.label_41.setText('暂无记录')
                self.label_41.repaint()
                self.label_471.setText('0')
                self.label_471.repaint()
            else:
                self.label_471.setText(str(len(self.jieru_result_page2)))
                self.label_471.repaint()
                if len(self.jieru_result_page2) <= 13:
                    for i in range(0, len(self.jieru_result_page2)):
                        self.label_projectname_page2[i].setText('   ' + self.jieru_result_page2[i][0])
                        self.label_partno_page2[i].setText('  ' + self.jieru_result_page2[i][1])
                        self.label_partname_page2[i].setText('      ' + self.jieru_result_page2[i][2])
                        self.label_barcode_page2[i].setText('   '+self.jieru_result_page2[i][3])
                        self.label_jierushijian_page2[i].setText(' ' + self.jieru_result_page2[i][4])
                        self.label_guazhangren_page2[i].setText('     ' + self.jieru_result_page2[i][5])
                else:  # 大于13
                    for i in range(0, 13):
                        self.label_projectname_page2[i].setText('   ' + self.jieru_result_page2[i][0])
                        self.label_partno_page2[i].setText('  ' + self.jieru_result_page2[i][1])
                        self.label_partname_page2[i].setText('      ' + self.jieru_result_page2[i][2])
                        self.label_barcode_page2[i].setText('   '+self.jieru_result_page2[i][3])
                        self.label_jierushijian_page2[i].setText(' ' + self.jieru_result_page2[i][4])
                        self.label_guazhangren_page2[i].setText('     ' + self.jieru_result_page2[i][5])
        elif self.lineEdit_4.text() == '' and self.lineEdit_3.text() != '':  # part name不为空
            page2_chaxun_partname_list = []  # 准备一个新的列表，筛选条件后，插入新的元素，依赖前面生成的所有的总的返回结果self.guazhang_result_page4，因为元组不能插入元素，写写列表后面转元组
            for x in range(0, 13):
                self.label_projectname_page2[x].setText('')
                self.label_partno_page2[x].setText('')
                self.label_partname_page2[x].setText('')
                self.label_barcode_page2[x].setText('')
                self.label_jierushijian_page2[x].setText('')
                self.label_guazhangren_page2[x].setText('')
            self.label_41.setText('')  # 88是显示暂无记录那个标签
            self.label_41.repaint()
            self.label_864.setText('1')
            self.label_864.repaint()
            # self.guazhang=(('LCH-2', '4X90S91831', 'CABLE USB-C TO ETHERNET ADP SC10P27625', '2022012111770'),('LCH-1', '20Y6Z9EBCN', 'THINKPAD X1 EXTREME GEN 4 SVT19 PRC', '2022012621453'))
            for each in self.jieru_result_page2:
                if self.lineEdit_3.text() in each[2]:  # 如果搜索的名字被包含在PARTname里
                    page2_chaxun_partname_list.append(each)
            self.jieru_result_page2 = tuple(page2_chaxun_partname_list)
            self.zongyeshu_page2, self.yushu_page2 = len(self.jieru_result_page2) // 13, len(self.jieru_result_page2) % 13  # //取整，%取余数, 后面会用
            if self.zongyeshu_page2 != 0 and self.yushu_page2 != 0:
                self.zongyeshu_page2 = self.zongyeshu_page2 + 1
            if self.zongyeshu_page2 == 0:
                self.label_862.setText('/1')
                self.label_862.repaint()
            else:
                self.label_862.setText(f'/{self.zongyeshu_page2}')
                self.label_862.repaint()
            if len(self.jieru_result_page2) == 0:
                self.label_41.setText('暂无记录')
                self.label_41.repaint()
                self.label_471.setText('0')
                self.label_471.repaint()
            else:
                self.label_471.setText(str(len(self.jieru_result_page2)))
                self.label_471.repaint()
                if len(self.jieru_result_page2) <= 13:
                    for i in range(0, len(self.jieru_result_page2)):
                        self.label_projectname_page2[i].setText('   ' + self.jieru_result_page2[i][0])
                        self.label_partno_page2[i].setText('  ' + self.jieru_result_page2[i][1])
                        self.label_partname_page2[i].setText('      ' + self.jieru_result_page2[i][2])
                        self.label_barcode_page2[i].setText('   '+self.jieru_result_page2[i][3])
                        self.label_jierushijian_page2[i].setText(' ' + self.jieru_result_page2[i][4])
                        self.label_guazhangren_page2[i].setText('     ' + self.jieru_result_page2[i][5])
                else:  # 大于13
                    for i in range(0, 13):
                        self.label_projectname_page2[i].setText('   ' + self.jieru_result_page2[i][0])
                        self.label_partno_page2[i].setText('  ' + self.jieru_result_page2[i][1])
                        self.label_partname_page2[i].setText('      ' + self.jieru_result_page2[i][2])
                        self.label_barcode_page2[i].setText('   '+self.jieru_result_page2[i][3])
                        self.label_jierushijian_page2[i].setText(' ' + self.jieru_result_page2[i][4])
                        self.label_guazhangren_page2[i].setText('     ' + self.jieru_result_page2[i][5])
        elif self.lineEdit_4.text() != '' and self.lineEdit_3.text() != '':  # part 和 project name不为空
            page2_chaxun_part_project_name_list = []  # 准备一个新的列表，筛选条件后，插入新的元素，依赖前面生成的所有的总的返回结果self.guazhang_result_page4，因为元组不能插入元素，写写列表后面转元组
            for x in range(0, 13):
                self.label_projectname_page2[x].setText('')
                self.label_partno_page2[x].setText('')
                self.label_partname_page2[x].setText('')
                self.label_barcode_page2[x].setText('')
                self.label_jierushijian_page2[x].setText('')
                self.label_guazhangren_page2[x].setText('')
            self.label_41.setText('')  # 88是显示暂无记录那个标签
            self.label_41.repaint()
            self.label_864.setText('1')
            self.label_864.repaint()
            # self.guazhang=(('LCH-2', '4X90S91831', 'CABLE USB-C TO ETHERNET ADP SC10P27625', '2022012111770'),('LCH-1', '20Y6Z9EBCN', 'THINKPAD X1 EXTREME GEN 4 SVT19 PRC', '2022012621453'))
            for each in self.jieru_result_page2:
                if self.lineEdit_3.text() in each[2] and self.lineEdit_4.text() in each[0]:  # 如果搜索的名字被包含在PARTname里
                    page2_chaxun_part_project_name_list.append(each)
            self.jieru_result_page2 = tuple(page2_chaxun_part_project_name_list)
            self.zongyeshu_page2, self.yushu_page2 = len(self.jieru_result_page2) // 13, len(self.jieru_result_page2) % 13  # //取整，%取余数, 后面会用
            if self.zongyeshu_page2 != 0 and self.yushu_page2 != 0:
                self.zongyeshu_page2 = self.zongyeshu_page2 + 1
            if self.zongyeshu_page2 == 0:
                self.label_862.setText('/1')
                self.label_862.repaint()
            else:
                self.label_862.setText(f'/{self.zongyeshu_page2}')
                self.label_862.repaint()
            if len(self.jieru_result_page2) == 0:
                self.label_41.setText('暂无记录')
                self.label_41.repaint()
                self.label_471.setText('0')
                self.label_471.repaint()
            else:
                self.label_471.setText(str(len(self.jieru_result_page2)))
                self.label_471.repaint()
                if len(self.jieru_result_page2) <= 13:
                    for i in range(0, len(self.jieru_result_page2)):
                        self.label_projectname_page2[i].setText('   ' + self.jieru_result_page2[i][0])
                        self.label_partno_page2[i].setText('  ' + self.jieru_result_page2[i][1])
                        self.label_partname_page2[i].setText('      ' + self.jieru_result_page2[i][2])
                        self.label_barcode_page2[i].setText('   '+self.jieru_result_page2[i][3])
                        self.label_jierushijian_page2[i].setText(' ' + self.jieru_result_page2[i][4])
                        self.label_guazhangren_page2[i].setText('     ' + self.jieru_result_page2[i][5])
                else:  # 大于13
                    for i in range(0, 13):
                        self.label_projectname_page2[i].setText('   ' + self.jieru_result_page2[i][0])
                        self.label_partno_page2[i].setText('  ' + self.jieru_result_page2[i][1])
                        self.label_partname_page2[i].setText('      ' + self.jieru_result_page2[i][2])
                        self.label_barcode_page2[i].setText('   '+self.jieru_result_page2[i][3])
                        self.label_jierushijian_page2[i].setText(' ' + self.jieru_result_page2[i][4])
                        self.label_guazhangren_page2[i].setText('     ' + self.jieru_result_page2[i][5])
        else:#如果2个都为空, 就什么都不做
            pass
    def chongzhi_page2(self):
        self.lineEdit_3.clear()
        self.lineEdit_4.clear()
    def xiayiye_page2(self):#只有大于13的时候，下一页button才有效，这里不用数据库，展示page2(self)的结果
        if int(self.label_864.text()) >=self.zongyeshu_page2:#如果当前页码大于等于总页码，点击不执行任何东西
            pass
        else:
            for x in range(0, 13):#点击下页的时候首先清掉当前页的所有内容，再展示下一页的.
                self.label_projectname_page2[x].setText('')
                self.label_partno_page2[x].setText('')
                self.label_partname_page2[x].setText('')
                self.label_barcode_page2[x].setText('')
                self.label_jierushijian_page2[x].setText('')
                self.label_guazhangren_page2[x].setText('')
            yema=int(self.label_864.text())
            self.label_864.setText(str(yema+1))
            self.label_864.repaint()
            current_yema = int(self.label_864.text())
            for y in range(0,13):
                try:
                    self.label_projectname_page2[y].setText('   ' + self.jieru_result_page2[(current_yema-1)*13+y][0])#y+13代表显示列表的13项开始，依次类推
                    self.label_projectname_page2[y].repaint()
                    self.label_partno_page2[y].setText('  ' + self.jieru_result_page2[(current_yema-1)*13+y][1])
                    self.label_partno_page2[y].repaint()
                    self.label_partname_page2[y].setText('      ' + self.jieru_result_page2[(current_yema-1)*13+y][2])
                    self.label_partname_page2[y].repaint()
                    self.label_barcode_page2[y].setText('   ' + self.jieru_result_page2[(current_yema-1)*13+y][3])
                    self.label_barcode_page2[y].repaint()
                    self.label_jierushijian_page2[y].setText(' ' + self.jieru_result_page2[(current_yema-1)*13+y][4])
                    self.label_jierushijian_page2[y].repaint()
                    self.label_guazhangren_page2[y].setText('     ' + self.jieru_result_page2[(current_yema-1)*13+y][5])
                    self.label_guazhangren_page2[y].repaint()
                except:###这里的作用是如果最后一页没有13个选项内容，前面肯定就会报错，这里就跳过报错，最后一页有多少就显示多少
                    pass
    def shangyiye_page2(self):  # 只有大于13的时候，下一页button才有效，这里不用数据库，展示page4(self)的结果
        if int(self.label_864.text()) ==1:  # 如果当前页码等于1的时候点击不执行任何动作
            pass
        else:
            for x in range(0, 13):  # 点击下页的时候首先清掉当前页的所有内容，再展示下一页的.
                self.label_projectname_page2[x].setText('')
                self.label_partno_page2[x].setText('')
                self.label_partname_page2[x].setText('')
                self.label_barcode_page2[x].setText('')
                self.label_jierushijian_page2[x].setText('')
                self.label_guazhangren_page2[x].setText('')
            yema = int(self.label_864.text())
            self.label_864.setText(str(yema - 1))
            self.label_864.repaint()
            current_yema = int(self.label_864.text())
            for y in range(0, 13):#上一页肯定会显示满，所以不需要用try
                self.label_projectname_page2[y].setText('   ' + self.jieru_result_page2[(current_yema - 1) * 13 + y][0])  # y+13代表显示列表的13项开始，依次类推
                self.label_projectname_page2[y].repaint()
                self.label_partno_page2[y].setText('  ' + self.jieru_result_page2[(current_yema - 1) * 13 + y][1])
                self.label_partno_page2[y].repaint()
                self.label_partname_page2[y].setText('      ' + self.jieru_result_page2[(current_yema - 1) * 13 + y][2])
                self.label_partname_page2[y].repaint()
                self.label_barcode_page2[y].setText('   ' + self.jieru_result_page2[(current_yema - 1) * 13 + y][3])
                self.label_barcode_page2[y].repaint()
                self.label_jierushijian_page2[y].setText(' ' + self.jieru_result_page2[(current_yema - 1) * 13 + y][4])
                self.label_jierushijian_page2[y].repaint()
                self.label_guazhangren_page2[y].setText('     ' + self.jieru_result_page2[(current_yema - 1) * 13 + y][5])
                self.label_guazhangren_page2[y].repaint()
    #上面是page2里包含的函数

    # 下面是page3里包含的函数
    def page3(self):
        self.lineEdit_5.clear()  # 清除条件查找框里的内容
        self.lineEdit_6.clear()  # 清除条件查找框里的内容
        for x in range(0, 13):
            self.label_projectname_page3[x].setText('')
            self.label_partname_page3[x].setText('')
            self.label_barcode_page3[x].setText('')
            self.label_guazhangren_page3[x].setText('')
            self.label_jieyongren_page3[x].setText('')
        self.label_69.setText('')#暂无记录那个label
        self.label_69.repaint()
        self.label_873.setText('1')
        self.label_873.repaint()
        self.stackedWidget.setCurrentIndex(2)
        try:
            db = pymysql.connect(host="localhost", user='root', password='jiang111', port=3306, database='lily', connect_timeout=3)
            cusor = db.cursor()
            sql = f"select * from {self.wuzichazhao}"
            cusor.execute(sql)
            self.wuzichazhao_result_page3 = cusor.fetchall()  # 返回这种(('LCH-2', '4X90S91831', 'CABLE USB-C TO ETHERNET ADP SC10P27625', '2022012111770'), ('LCH-1', '20Y6Z9EBCN', 'THINKPAD X1 EXTREME GEN 4 SVT19 PRC', '2022012621453'))
            self.yuanshi_shuju_page3 = self.wuzichazhao_result_page3  # 这里是为了后面每一次条件查询前先恢复到原始data
            self.zongyeshu_page3, self.yushu_page3 = len(self.wuzichazhao_result_page3) // 13, len(self.wuzichazhao_result_page3) % 13  # //取整，%取余数, 后面会用
            if self.zongyeshu_page3 != 0 and self.yushu_page3 != 0:
                self.zongyeshu_page3 = self.zongyeshu_page3 + 1
            if self.zongyeshu_page3 == 0:
                self.label_871.setText('/1')
                self.label_871.repaint()
            else:
                self.label_871.setText(f'/{self.zongyeshu_page3}')
                self.label_871.repaint()
            if len(self.wuzichazhao_result_page3) == 0:
                self.label_69.setText('暂无记录')
                self.label_69.repaint()
                self.label_652.setText('0')
                self.label_652.repaint()
            else:
                self.label_652.setText(str(len(self.wuzichazhao_result_page3)))
                self.label_652.repaint()
                if len(self.wuzichazhao_result_page3) <= 13:
                    for i in range(0, len(self.wuzichazhao_result_page3)):
                        self.label_projectname_page3[i].setText('    ' + self.wuzichazhao_result_page3[i][0])
                        self.label_partname_page3[i].setText('         ' + self.wuzichazhao_result_page3[i][1])
                        self.label_barcode_page3[i].setText('   '+self.wuzichazhao_result_page3[i][2])
                        self.label_guazhangren_page3[i].setText('  ' + self.wuzichazhao_result_page3[i][3])
                        self.label_jieyongren_page3[i].setText('  ' + self.wuzichazhao_result_page3[i][4])
                else:  # 大于13
                    for i in range(0, 13):
                        self.label_projectname_page3[i].setText('    ' + self.wuzichazhao_result_page3[i][0])
                        self.label_partname_page3[i].setText('         ' + self.wuzichazhao_result_page3[i][1])
                        self.label_barcode_page3[i].setText('   '+self.wuzichazhao_result_page3[i][2])
                        self.label_guazhangren_page3[i].setText('  ' + self.wuzichazhao_result_page3[i][3])
                        self.label_jieyongren_page3[i].setText('  ' + self.wuzichazhao_result_page3[i][4])
                db.close()
        except:
            self.label_69.setText('暂无记录')
            self.label_69.repaint()
            self.label_652.setText('0')#0条记录
            self.label_652.repaint()
    def chaxun_page3(self):#查询只设置3种情况，要么2个标签都有内容，要么其中一个有，否则查询函数为pass
        self.wuzichazhao_result_page3=self.yuanshi_shuju_page3#每次复原后再查询
        if self.lineEdit_5.text()!='' and self.lineEdit_6.text()=='':#project name不为空
            page3_chaxun_project_list=[]#准备一个新的列表，筛选条件后，插入新的元素，依赖前面生成的所有的总的返回结果self.guazhang_result_page4，因为元组不能插入元素，写写列表后面转元组
            for x in range(0, 13):
                self.label_projectname_page3[x].setText('')
                self.label_partname_page3[x].setText('')
                self.label_barcode_page3[x].setText('')
                self.label_guazhangren_page3[x].setText('')
                self.label_jieyongren_page3[x].setText('')
            self.label_69.setText('')  # 暂无记录那个label
            self.label_69.repaint()
            self.label_873.setText('1')
            self.label_873.repaint()
            #self.guazhang=(('LCH-2', '4X90S91831', 'CABLE USB-C TO ETHERNET ADP SC10P27625', '2022012111770'),('LCH-1', '20Y6Z9EBCN', 'THINKPAD X1 EXTREME GEN 4 SVT19 PRC', '2022012621453'))
            for each in self.wuzichazhao_result_page3:
                if self.lineEdit_5.text() in each[0]:#如果搜索的名字被包含在project name里
                    page3_chaxun_project_list.append(each)
            self.wuzichazhao_result_page3=tuple(page3_chaxun_project_list)
            self.zongyeshu_page3, self.yushu_page3 = len(self.wuzichazhao_result_page3) // 13, len(self.wuzichazhao_result_page3) % 13  # //取整，%取余数, 后面会用
            if self.zongyeshu_page3 != 0 and self.yushu_page3 != 0:
                self.zongyeshu_page3 = self.zongyeshu_page3 + 1
            if self.zongyeshu_page3 == 0:
                self.label_871.setText('/1')
                self.label_871.repaint()
            else:
                self.label_871.setText(f'/{self.zongyeshu_page3}')
                self.label_871.repaint()
            if len(self.wuzichazhao_result_page3) == 0:
                self.label_69.setText('暂无记录')
                self.label_69.repaint()
                self.label_652.setText('0')
                self.label_652.repaint()
            else:
                self.label_652.setText(str(len(self.wuzichazhao_result_page3)))
                self.label_652.repaint()
                if len(self.wuzichazhao_result_page3) <= 13:
                    for i in range(0, len(self.wuzichazhao_result_page3)):
                        self.label_projectname_page3[i].setText('    ' + self.wuzichazhao_result_page3[i][0])
                        self.label_partname_page3[i].setText('         ' + self.wuzichazhao_result_page3[i][1])
                        self.label_barcode_page3[i].setText('   '+self.wuzichazhao_result_page3[i][2])
                        self.label_guazhangren_page3[i].setText('  ' + self.wuzichazhao_result_page3[i][3])
                        self.label_jieyongren_page3[i].setText('  ' + self.wuzichazhao_result_page3[i][4])
                else:  # 大于13
                    for i in range(0, 13):
                        self.label_projectname_page3[i].setText('    ' + self.wuzichazhao_result_page3[i][0])
                        self.label_partname_page3[i].setText('         ' + self.wuzichazhao_result_page3[i][1])
                        self.label_barcode_page3[i].setText('   '+self.wuzichazhao_result_page3[i][2])
                        self.label_guazhangren_page3[i].setText('  ' + self.wuzichazhao_result_page3[i][3])
                        self.label_jieyongren_page3[i].setText('  ' + self.wuzichazhao_result_page3[i][4])
        elif self.lineEdit_5.text() == '' and self.lineEdit_6.text() != '':  # part name不为空
            page3_chaxun_partname_list = []  # 准备一个新的列表，筛选条件后，插入新的元素，依赖前面生成的所有的总的返回结果self.guazhang_result_page4，因为元组不能插入元素，写写列表后面转元组
            for x in range(0, 13):
                self.label_projectname_page3[x].setText('')
                self.label_partname_page3[x].setText('')
                self.label_barcode_page3[x].setText('')
                self.label_guazhangren_page3[x].setText('')
                self.label_jieyongren_page3[x].setText('')
            self.label_69.setText('')  # 暂无记录那个label
            self.label_69.repaint()
            self.label_873.setText('1')
            self.label_873.repaint()
            # self.guazhang=(('LCH-2', '4X90S91831', 'CABLE USB-C TO ETHERNET ADP SC10P27625', '2022012111770'),('LCH-1', '20Y6Z9EBCN', 'THINKPAD X1 EXTREME GEN 4 SVT19 PRC', '2022012621453'))
            for each in self.wuzichazhao_result_page3:
                if self.lineEdit_6  .text() in each[1]:  # 如果搜索的名字被包含在PARTname里
                    page3_chaxun_partname_list.append(each)
            self.wuzichazhao_result_page3 = tuple(page3_chaxun_partname_list)
            self.zongyeshu_page3, self.yushu_page3 = len(self.wuzichazhao_result_page3) // 13, len(self.wuzichazhao_result_page3) % 13  # //取整，%取余数, 后面会用
            if self.zongyeshu_page3 != 0 and self.yushu_page3 != 0:
                self.zongyeshu_page3 = self.zongyeshu_page3 + 1
            if self.zongyeshu_page3 == 0:
                self.label_871.setText('/1')
                self.label_871.repaint()
            else:
                self.label_871.setText(f'/{self.zongyeshu_page3}')
                self.label_871.repaint()
            if len(self.wuzichazhao_result_page3) == 0:
                self.label_69.setText('暂无记录')
                self.label_69.repaint()
                self.label_652.setText('0')
                self.label_652.repaint()
            else:
                self.label_652.setText(str(len(self.wuzichazhao_result_page3)))
                self.label_652.repaint()
                if len(self.wuzichazhao_result_page3) <= 13:
                    for i in range(0, len(self.wuzichazhao_result_page3)):
                        self.label_projectname_page3[i].setText('    ' + self.wuzichazhao_result_page3[i][0])
                        self.label_partname_page3[i].setText('         ' + self.wuzichazhao_result_page3[i][1])
                        self.label_barcode_page3[i].setText('   '+self.wuzichazhao_result_page3[i][2])
                        self.label_guazhangren_page3[i].setText('  ' + self.wuzichazhao_result_page3[i][3])
                        self.label_jieyongren_page3[i].setText('  ' + self.wuzichazhao_result_page3[i][4])
                else:  # 大于13
                    for i in range(0, 13):
                        self.label_projectname_page3[i].setText('    ' + self.wuzichazhao_result_page3[i][0])
                        self.label_partname_page3[i].setText('         ' + self.wuzichazhao_result_page3[i][1])
                        self.label_barcode_page3[i].setText('   '+self.wuzichazhao_result_page3[i][2])
                        self.label_guazhangren_page3[i].setText('  ' + self.wuzichazhao_result_page3[i][3])
                        self.label_jieyongren_page3[i].setText('  ' + self.wuzichazhao_result_page3[i][4])
        elif self.lineEdit_5.text() != '' and self.lineEdit_6.text() != '':  # part 和 project name不为空
            page3_chaxun_part_project_name_list = []  # 准备一个新的列表，筛选条件后，插入新的元素，依赖前面生成的所有的总的返回结果self.guazhang_result_page4，因为元组不能插入元素，写写列表后面转元组
            for x in range(0, 13):
                self.label_projectname_page3[x].setText('')
                self.label_partname_page3[x].setText('')
                self.label_barcode_page3[x].setText('')
                self.label_guazhangren_page3[x].setText('')
                self.label_jieyongren_page3[x].setText('')
            self.label_69.setText('')  # 暂无记录那个label
            self.label_69.repaint()
            self.label_873.setText('1')
            self.label_873.repaint()
            # self.guazhang=(('LCH-2', '4X90S91831', 'CABLE USB-C TO ETHERNET ADP SC10P27625', '2022012111770'),('LCH-1', '20Y6Z9EBCN', 'THINKPAD X1 EXTREME GEN 4 SVT19 PRC', '2022012621453'))
            for each in self.wuzichazhao_result_page3:
                if self.lineEdit_5.text() in each[0] and self.lineEdit_6.text() in each[1]:  # 如果搜索的名字被包含在PARTname里
                    page3_chaxun_part_project_name_list.append(each)
            self.wuzichazhao_result_page3 = tuple(page3_chaxun_part_project_name_list)
            self.zongyeshu_page3, self.yushu_page3 = len(self.wuzichazhao_result_page3) // 13, len(self.wuzichazhao_result_page3) % 13  # //取整，%取余数, 后面会用
            if self.zongyeshu_page3 != 0 and self.yushu_page3 != 0:
                self.zongyeshu_page3 = self.zongyeshu_page3 + 1
            if self.zongyeshu_page3 == 0:
                self.label_871.setText('/1')
                self.label_871.repaint()
            else:
                self.label_871.setText(f'/{self.zongyeshu_page3}')
                self.label_871.repaint()
            if len(self.wuzichazhao_result_page3) == 0:
                self.label_69.setText('暂无记录')
                self.label_69.repaint()
                self.label_652.setText('0')
                self.label_652.repaint()
            else:
                self.label_652.setText(str(len(self.wuzichazhao_result_page3)))
                self.label_652.repaint()
                if len(self.wuzichazhao_result_page3) <= 13:
                    for i in range(0, len(self.wuzichazhao_result_page3)):
                        self.label_projectname_page3[i].setText('    ' + self.wuzichazhao_result_page3[i][0])
                        self.label_partname_page3[i].setText('         ' + self.wuzichazhao_result_page3[i][1])
                        self.label_barcode_page3[i].setText('   '+self.wuzichazhao_result_page3[i][2])
                        self.label_guazhangren_page3[i].setText('  ' + self.wuzichazhao_result_page3[i][3])
                        self.label_jieyongren_page3[i].setText('  ' + self.wuzichazhao_result_page3[i][4])
                else:  # 大于13
                    for i in range(0, 13):
                        self.label_projectname_page3[i].setText('    ' + self.wuzichazhao_result_page3[i][0])
                        self.label_partname_page3[i].setText('         ' + self.wuzichazhao_result_page3[i][1])
                        self.label_barcode_page3[i].setText('   '+self.wuzichazhao_result_page3[i][2])
                        self.label_guazhangren_page3[i].setText('  ' + self.wuzichazhao_result_page3[i][3])
                        self.label_jieyongren_page3[i].setText('  ' + self.wuzichazhao_result_page3[i][4])
        else:#如果2个都为空, 就什么都不做
            pass
    def chongzhi_page3(self):
        self.lineEdit_5.clear()
        self.lineEdit_6.clear()
    def xiayiye_page3(self):#只有大于13的时候，下一页button才有效，这里不用数据库，展示page2(self)的结果
        if int(self.label_873.text()) >=self.zongyeshu_page3:#如果当前页码大于等于总页码，点击不执行任何东西
            pass
        else:
            for x in range(0, 13):#点击下页的时候首先清掉当前页的所有内容，再展示下一页的.
                self.label_projectname_page3[x].setText('')
                self.label_partname_page3[x].setText('')
                self.label_barcode_page3[x].setText('')
                self.label_guazhangren_page3[x].setText('')
                self.label_jieyongren_page3[x].setText('')
            yema=int(self.label_873.text())
            self.label_873.setText(str(yema+1))
            self.label_873.repaint()
            current_yema = int(self.label_873.text())
            for y in range(0,13):
                try:
                    self.label_projectname_page3[y].setText('    ' + self.wuzichazhao_result_page3[(current_yema-1)*13+y][0])#y+13代表显示列表的13项开始，依次类推
                    self.label_projectname_page3[y].repaint()
                    self.label_partname_page3[y].setText('         ' + self.wuzichazhao_result_page3[(current_yema-1)*13+y][1])
                    self.label_partname_page3[y].repaint()
                    self.label_barcode_page3[y].setText('   ' + self.wuzichazhao_result_page3[(current_yema-1)*13+y][2])
                    self.label_barcode_page3[y].repaint()
                    self.label_guazhangren_page3[y].setText('  ' + self.wuzichazhao_result_page3[(current_yema-1)*13+y][3])
                    self.label_guazhangren_page3[y].repaint()
                    self.label_jieyongren_page3[y].setText('  ' + self.wuzichazhao_result_page3[(current_yema-1)*13+y][4])
                    self.label_jieyongren_page3[y].repaint()
                except:###这里的作用是如果最后一页没有13个选项内容，前面肯定就会报错，这里就跳过报错，最后一页有多少就显示多少
                    pass
    def shangyiye_page3(self):  # 只有大于13的时候，下一页button才有效，这里不用数据库，展示page4(self)的结果
        if int(self.label_873.text()) ==1:  # 如果当前页码等于1的时候点击不执行任何动作
            pass
        else:
            for x in range(0, 13):  # 点击下页的时候首先清掉当前页的所有内容，再展示下一页的.
                self.label_projectname_page3[x].setText('')
                self.label_partname_page3[x].setText('')
                self.label_barcode_page3[x].setText('')
                self.label_guazhangren_page3[x].setText('')
                self.label_jieyongren_page3[x].setText('')
            yema = int(self.label_873.text())
            self.label_873.setText(str(yema - 1))
            self.label_873.repaint()
            current_yema = int(self.label_873.text())
            for y in range(0, 13):#上一页肯定会显示满，所以不需要用try
                self.label_projectname_page3[y].setText('    ' + self.wuzichazhao_result_page3[(current_yema - 1) * 13 + y][0])  # y+13代表显示列表的13项开始，依次类推
                self.label_projectname_page3[y].repaint()
                self.label_partname_page3[y].setText('         ' + self.wuzichazhao_result_page3[(current_yema - 1) * 13 + y][1])
                self.label_partname_page3[y].repaint()
                self.label_barcode_page3[y].setText('   ' + self.wuzichazhao_result_page3[(current_yema - 1) * 13 + y][2])
                self.label_barcode_page3[y].repaint()
                self.label_guazhangren_page3[y].setText('  ' + self.wuzichazhao_result_page3[(current_yema - 1) * 13 + y][3])
                self.label_guazhangren_page3[y].repaint()
                self.label_jieyongren_page3[y].setText('  ' + self.wuzichazhao_result_page3[(current_yema - 1) * 13 + y][4])
                self.label_jieyongren_page3[y].repaint()
    # 上面是page3里包含的函数

    #下面是page4里包含的函数
    def page4(self):
        self.frame_3.hide()
        self.lineEdit_7.clear()#清除条件查找框里的内容
        self.lineEdit_8.clear()#清除条件查找框里的内容
        for x in range(0, 13):
            self.label_projectname_page4[x].setText('')
            self.label_partno_page4[x].setText('')
            self.label_partname_page4[x].setText('')
            self.label_barcode_page4[x].setText('')
            self.label_satus_page4[x].setText('')
            self.jiechu_button_page4[x].setText('')
        self.label_88.setText('')#暂无记录那个label
        self.label_88.repaint()
        self.label_220.setText('1')
        self.label_220.repaint()
        self.stackedWidget.setCurrentIndex(3)
        try:
            db = pymysql.connect(host="localhost", user='root', password='jiang111', port=3306, database='lily', connect_timeout=3)
            cusor = db.cursor()
            self.table_str = 'guazhang_' + user
            sql=f"select * from {self.table_str}"
            cusor.execute(sql)
            self.guazhang_result_page4=cusor.fetchall()#返回这种(('LCH-2', '4X90S91831', 'CABLE USB-C TO ETHERNET ADP SC10P27625', '2022012111770'), ('LCH-1', '20Y6Z9EBCN', 'THINKPAD X1 EXTREME GEN 4 SVT19 PRC', '2022012621453'))
            self.yuanshi_shuju=self.guazhang_result_page4#这里是为了后面每一次条件查询前先恢复到原始data
            self.zongyeshu, self.yushu = len(self.guazhang_result_page4) // 13, len(self.guazhang_result_page4) % 13  # //取整，%取余数, 后面会用
            if self.zongyeshu != 0 and self.yushu != 0:
                self.zongyeshu=self.zongyeshu+1
            if self.zongyeshu==0:
                self.label_228.setText('/1')
                self.label_228.repaint()
            else:
                self.label_228.setText(f'/{self.zongyeshu}')
                self.label_228.repaint()
            if len(self.guazhang_result_page4)==0:
                self.label_88.setText('暂无记录')
                self.label_88.repaint()
                self.label_112.setText('0')
                self.label_112.repaint()
            else:
                self.label_112.setText(str(len(self.guazhang_result_page4)))
                self.label_112.repaint()
                if len(self.guazhang_result_page4) <=13:
                    for i in range(0,len(self.guazhang_result_page4)):
                        self.label_projectname_page4[i].setText(' '+self.guazhang_result_page4[i][0])
                        self.label_partno_page4[i].setText('  '+self.guazhang_result_page4[i][1])
                        self.label_partname_page4[i].setText(' '+self.guazhang_result_page4[i][2])
                        self.label_barcode_page4[i].setText(self.guazhang_result_page4[i][3])
                        self.label_satus_page4[i].setText('   '+self.guazhang_result_page4[i][4])
                        self.jiechu_button_page4[i].setText('                                     借出')
                        if self.guazhang_result_page4[i][4] == '已借出':
                            self.jiechu_button_page4[i].setEnabled(False)
                        else:
                            self.jiechu_button_page4[i].setEnabled(True)
                else:#大于13
                    for i in range(0,13):
                        self.label_projectname_page4[i].setText(' ' + self.guazhang_result_page4[i][0])
                        self.label_partno_page4[i].setText('  ' + self.guazhang_result_page4[i][1])
                        self.label_partname_page4[i].setText(' ' + self.guazhang_result_page4[i][2])
                        self.label_barcode_page4[i].setText(self.guazhang_result_page4[i][3])
                        self.label_satus_page4[i].setText('   ' + self.guazhang_result_page4[i][4])
                        self.jiechu_button_page4[i].setText('                                     借出')
                        if self.guazhang_result_page4[i][4] == '已借出':
                            self.jiechu_button_page4[i].setEnabled(False)
                        else:
                            self.jiechu_button_page4[i].setEnabled(True)
                db.close()
        except:
            self.label_88.setText('暂无记录')
            self.label_88.repaint()
            self.label_112.setText('0')
            self.label_112.repaint()
        list1_page4=[self.label_92.text(),self.label_120.text(),self.label_121.text(),self.label_122.text(),self.label_123.text()]#第一行对应的label的各项内容project name 等
        list2_page4 = [self.label_93.text(), self.label_130.text(), self.label_129.text(), self.label_131.text(), self.label_127.text()]
        list3_page4 = [self.label_94.text(), self.label_137.text(), self.label_136.text(), self.label_138.text(), self.label_134.text()]  # 第一行对应的label的各项内容project name 等
        list4_page4 = [self.label_95.text(), self.label_145.text(), self.label_143.text(), self.label_141.text(), self.label_144.text()]
        list5_page4 = [self.label_98.text(), self.label_148.text(), self.label_150.text(), self.label_149.text(), self.label_153.text()]
        list6_page4 = [self.label_99.text(), self.label_156.text(), self.label_155.text(), self.label_160.text(), self.label_157.text()]
        list7_page4 = [self.label_166.text(), self.label_168.text(), self.label_162.text(), self.label_164.text(), self.label_165.text()]
        list8_page4 = [self.label_175.text(), self.label_177.text(), self.label_173.text(), self.label_170.text(), self.label_171.text()]
        list9_page4 = [self.label_178.text(), self.label_179.text(), self.label_181.text(), self.label_184.text(), self.label_185.text()]
        list10_page4 = [self.label_188.text(), self.label_187.text(), self.label_186.text(), self.label_191.text(), self.label_192.text()]
        list11_page4 = [self.label_194.text(), self.label_199.text(), self.label_197.text(), self.label_196.text(), self.label_200.text()]
        list12_page4 = [self.label_209.text(), self.label_208.text(), self.label_204.text(), self.label_207.text(), self.label_203.text()]
        list13_page4 = [self.label_226.text(), self.label_227.text(), self.label_214.text(), self.label_225.text(), self.label_213.text()]
        self.jiechu_button_duiying_label_page4={'pushButton_24':list1_page4,'pushButton_47':list2_page4,'pushButton_26':list3_page4,'pushButton_48':list4_page4,'pushButton_27':list5_page4,'pushButton_49':list6_page4,'pushButton_29':list7_page4,'pushButton_50':list8_page4,'pushButton_30':list9_page4,'pushButton_51':list10_page4,'pushButton_31':list11_page4,'pushButton_52':list12_page4,'pushButton_33':list13_page4}
    def xiayiye_page4(self):#只有大于13的时候，下一页button才有效，这里不用数据库，展示page4(self)的结果
        if int(self.label_220.text()) >=self.zongyeshu:#如果当前页码大于等于总页码，点击不执行任何东西
            pass
        else:
            for x in range(0, 13):#点击下页的时候首先清掉当前页的所有内容，再展示下一页的.
                self.label_projectname_page4[x].setText('')
                self.label_partno_page4[x].setText('')
                self.label_partname_page4[x].setText('')
                self.label_barcode_page4[x].setText('')
                self.label_satus_page4[x].setText('')
                self.jiechu_button_page4[x].setText('')
            yema=int(self.label_220.text())
            self.label_220.setText(str(yema+1))
            self.label_220.repaint()
            current_yema = int(self.label_220.text())
            for y in range(0,13):
                try:
                    self.label_projectname_page4[y].setText(' ' + self.guazhang_result_page4[(current_yema-1)*13+y][0])#y+13代表显示列表的13项开始，依次类推
                    self.label_projectname_page4[y].repaint()
                    self.label_partno_page4[y].setText('  ' + self.guazhang_result_page4[(current_yema-1)*13+y][1])
                    self.label_partno_page4[y].repaint()
                    self.label_partname_page4[y].setText(' ' + self.guazhang_result_page4[(current_yema-1)*13+y][2])
                    self.label_partname_page4[y].repaint()
                    self.label_barcode_page4[y].setText(self.guazhang_result_page4[(current_yema-1)*13+y][3])
                    self.label_barcode_page4[y].repaint()
                    self.label_satus_page4[y].setText('   ' + self.guazhang_result_page4[(current_yema-1)*13+y][4])
                    self.label_satus_page4[y].repaint()
                    self.jiechu_button_page4[y].setText('                                     借出')
                    if self.guazhang_result_page4[(current_yema-1)*13+y][4] == '已借出':
                        self.jiechu_button_page4[y].setEnabled(False)
                    else:
                        self.jiechu_button_page4[y].setEnabled(True)

                except:###这里的作用是如果最后一页没有13个选项内容，前面肯定就会报错，这里就跳过报错，最后一页有多少就显示多少
                    pass
        list1_page4 = [self.label_92.text(), self.label_120.text(), self.label_121.text(), self.label_122.text(), self.label_123.text()]  # 第一行对应的label的各项内容project name 等
        list2_page4 = [self.label_93.text(), self.label_130.text(), self.label_129.text(), self.label_131.text(), self.label_127.text()]
        list3_page4 = [self.label_94.text(), self.label_137.text(), self.label_136.text(), self.label_138.text(), self.label_134.text()]  # 第一行对应的label的各项内容project name 等
        list4_page4 = [self.label_95.text(), self.label_145.text(), self.label_143.text(), self.label_141.text(), self.label_144.text()]
        list5_page4 = [self.label_98.text(), self.label_148.text(), self.label_150.text(), self.label_149.text(), self.label_153.text()]
        list6_page4 = [self.label_99.text(), self.label_156.text(), self.label_155.text(), self.label_160.text(), self.label_157.text()]
        list7_page4 = [self.label_166.text(), self.label_168.text(), self.label_162.text(), self.label_164.text(), self.label_165.text()]
        list8_page4 = [self.label_175.text(), self.label_177.text(), self.label_173.text(), self.label_170.text(), self.label_171.text()]
        list9_page4 = [self.label_178.text(), self.label_179.text(), self.label_181.text(), self.label_184.text(), self.label_185.text()]
        list10_page4 = [self.label_188.text(), self.label_187.text(), self.label_186.text(), self.label_191.text(), self.label_192.text()]
        list11_page4 = [self.label_194.text(), self.label_199.text(), self.label_197.text(), self.label_196.text(), self.label_200.text()]
        list12_page4 = [self.label_209.text(), self.label_208.text(), self.label_204.text(), self.label_207.text(), self.label_203.text()]
        list13_page4 = [self.label_226.text(), self.label_227.text(), self.label_214.text(), self.label_225.text(), self.label_213.text()]
        self.jiechu_button_duiying_label_page4 = {'pushButton_24': list1_page4, 'pushButton_47': list2_page4, 'pushButton_26': list3_page4, 'pushButton_48': list4_page4, 'pushButton_27': list5_page4, 'pushButton_49': list6_page4, 'pushButton_29': list7_page4, 'pushButton_50': list8_page4, 'pushButton_30': list9_page4, 'pushButton_51': list10_page4, 'pushButton_31': list11_page4, 'pushButton_52': list12_page4, 'pushButton_33': list13_page4}
    def shangyiye_page4(self):  # 只有大于13的时候，下一页button才有效，这里不用数据库，展示page4(self)的结果
        if int(self.label_220.text()) ==1:  # 如果当前页码等于1的时候点击不执行任何动作
            pass
        else:
            for x in range(0, 13):  # 点击下页的时候首先清掉当前页的所有内容，再展示下一页的.
                self.label_projectname_page4[x].setText('')
                self.label_partno_page4[x].setText('')
                self.label_partname_page4[x].setText('')
                self.label_barcode_page4[x].setText('')
                self.label_satus_page4[x].setText('')
                self.jiechu_button_page4[x].setText('')
            yema = int(self.label_220.text())
            self.label_220.setText(str(yema - 1))
            self.label_220.repaint()
            current_yema = int(self.label_220.text())
            for y in range(0, 13):#上一页肯定会显示满，所以不需要用try
                self.label_projectname_page4[y].setText(' ' + self.guazhang_result_page4[(current_yema - 1) * 13 + y][0])  # y+13代表显示列表的13项开始，依次类推
                self.label_projectname_page4[y].repaint()
                self.label_partno_page4[y].setText('  ' + self.guazhang_result_page4[(current_yema - 1) * 13 + y][1])
                self.label_partno_page4[y].repaint()
                self.label_partname_page4[y].setText(' ' + self.guazhang_result_page4[(current_yema - 1) * 13 + y][2])
                self.label_partname_page4[y].repaint()
                self.label_barcode_page4[y].setText(self.guazhang_result_page4[(current_yema - 1) * 13 + y][3])
                self.label_barcode_page4[y].repaint()
                self.label_satus_page4[y].setText('   ' + self.guazhang_result_page4[(current_yema - 1) * 13 + y][4])
                self.label_satus_page4[y].repaint()
                self.jiechu_button_page4[y].setText('                                     借出')
                if self.guazhang_result_page4[(current_yema - 1) * 13 + y][4] == '已借出':
                  self.jiechu_button_page4[y].setEnabled(False)
                else:
                  self.jiechu_button_page4[y].setEnabled(True)
        list1_page4=[self.label_92.text(),self.label_120.text(),self.label_121.text(),self.label_122.text(),self.label_123.text()]#第一行对应的label的各项内容project name 等
        list2_page4 = [self.label_93.text(), self.label_130.text(), self.label_129.text(), self.label_131.text(), self.label_127.text()]
        list3_page4 = [self.label_94.text(), self.label_137.text(), self.label_136.text(), self.label_138.text(), self.label_134.text()]  # 第一行对应的label的各项内容project name 等
        list4_page4 = [self.label_95.text(), self.label_145.text(), self.label_143.text(), self.label_141.text(), self.label_144.text()]
        list5_page4 = [self.label_98.text(), self.label_148.text(), self.label_150.text(), self.label_149.text(), self.label_153.text()]
        list6_page4 = [self.label_99.text(), self.label_156.text(), self.label_155.text(), self.label_160.text(), self.label_157.text()]
        list7_page4 = [self.label_166.text(), self.label_168.text(), self.label_162.text(), self.label_164.text(), self.label_165.text()]
        list8_page4 = [self.label_175.text(), self.label_177.text(), self.label_173.text(), self.label_170.text(), self.label_171.text()]
        list9_page4 = [self.label_178.text(), self.label_179.text(), self.label_181.text(), self.label_184.text(), self.label_185.text()]
        list10_page4 = [self.label_188.text(), self.label_187.text(), self.label_186.text(), self.label_191.text(), self.label_192.text()]
        list11_page4 = [self.label_194.text(), self.label_199.text(), self.label_197.text(), self.label_196.text(), self.label_200.text()]
        list12_page4 = [self.label_209.text(), self.label_208.text(), self.label_204.text(), self.label_207.text(), self.label_203.text()]
        list13_page4 = [self.label_226.text(), self.label_227.text(), self.label_214.text(), self.label_225.text(), self.label_213.text()]
        self.jiechu_button_duiying_label_page4={'pushButton_24':list1_page4,'pushButton_47':list2_page4,'pushButton_26':list3_page4,'pushButton_48':list4_page4,'pushButton_27':list5_page4,'pushButton_49':list6_page4,'pushButton_29':list7_page4,'pushButton_50':list8_page4,'pushButton_30':list9_page4,'pushButton_51':list10_page4,'pushButton_31':list11_page4,'pushButton_52':list12_page4,'pushButton_33':list13_page4}
    def chongzhi_page4(self):
        self.lineEdit_7.clear()
        self.lineEdit_8.clear()
    def chaxun_page4(self):#查询只设置3种情况，要么2个标签都有内容，要么其中一个有，否则查询函数为pass
        self.guazhang_result_page4=self.yuanshi_shuju#每次复原后再查询
        if self.lineEdit_7.text()!='' and self.lineEdit_8.text()=='':#project name不为空
            page4_chaxun_project_list=[]#准备一个新的列表，筛选条件后，插入新的元素，依赖前面生成的所有的总的返回结果self.guazhang_result_page4，因为元组不能插入元素，写写列表后面转元组
            for x in range(0, 13):
                self.label_projectname_page4[x].setText('')
                self.label_partno_page4[x].setText('')
                self.label_partname_page4[x].setText('')
                self.label_barcode_page4[x].setText('')
                self.label_satus_page4[x].setText('')
                self.jiechu_button_page4[x].setText('')
            self.label_88.setText('')#88是显示暂无记录那个标签
            self.label_88.repaint()
            self.label_220.setText('1')
            self.label_220.repaint()
            #self.guazhang=(('LCH-2', '4X90S91831', 'CABLE USB-C TO ETHERNET ADP SC10P27625', '2022012111770'),('LCH-1', '20Y6Z9EBCN', 'THINKPAD X1 EXTREME GEN 4 SVT19 PRC', '2022012621453'))
            for each in self.guazhang_result_page4:
                if self.lineEdit_7.text() in each[0]:#如果搜索的名字被包含在project name里
                    page4_chaxun_project_list.append(each)
            self.guazhang_result_page4=tuple(page4_chaxun_project_list)
            self.zongyeshu, self.yushu = len(self.guazhang_result_page4) // 13, len(self.guazhang_result_page4) % 13  # //取整，%取余数, 后面会用
            if self.zongyeshu != 0 and self.yushu != 0:
                self.zongyeshu = self.zongyeshu + 1
            if self.zongyeshu == 0:
                self.label_228.setText('/1')
                self.label_228.repaint()
            else:
                self.label_228.setText(f'/{self.zongyeshu}')
                self.label_228.repaint()
            if len(self.guazhang_result_page4) == 0:
                self.label_88.setText('暂无记录')
                self.label_88.repaint()
                self.label_112.setText('0')
                self.label_112.repaint()
            else:
                self.label_112.setText(str(len(self.guazhang_result_page4)))
                self.label_112.repaint()
                if len(self.guazhang_result_page4) <= 13:
                    for i in range(0, len(self.guazhang_result_page4)):
                        self.label_projectname_page4[i].setText(' ' + self.guazhang_result_page4[i][0])
                        self.label_partno_page4[i].setText('  ' + self.guazhang_result_page4[i][1])
                        self.label_partname_page4[i].setText(' ' + self.guazhang_result_page4[i][2])
                        self.label_barcode_page4[i].setText(self.guazhang_result_page4[i][3])
                        self.label_satus_page4[i].setText('   ' + self.guazhang_result_page4[i][4])
                        self.jiechu_button_page4[i].setText('                                     借出')
                        if self.guazhang_result_page4[i][4] == '已借出':
                            self.jiechu_button_page4[i].setEnabled(False)
                        else:
                            self.jiechu_button_page4[i].setEnabled(True)
                else:  # 大于13
                    for i in range(0, 13):
                        self.label_projectname_page4[i].setText(' ' + self.guazhang_result_page4[i][0])
                        self.label_partno_page4[i].setText('  ' + self.guazhang_result_page4[i][1])
                        self.label_partname_page4[i].setText(' ' + self.guazhang_result_page4[i][2])
                        self.label_barcode_page4[i].setText(self.guazhang_result_page4[i][3])
                        self.label_satus_page4[i].setText('   ' + self.guazhang_result_page4[i][4])
                        self.jiechu_button_page4[i].setText('                                     借出')
                        if self.guazhang_result_page4[i][4] == '已借出':
                            self.jiechu_button_page4[i].setEnabled(False)
                        else:
                            self.jiechu_button_page4[i].setEnabled(True)
        elif self.lineEdit_7.text() == '' and self.lineEdit_8.text() != '':  # part name不为空
            page4_chaxun_partname_list = []  # 准备一个新的列表，筛选条件后，插入新的元素，依赖前面生成的所有的总的返回结果self.guazhang_result_page4，因为元组不能插入元素，写写列表后面转元组
            for x in range(0, 13):
                self.label_projectname_page4[x].setText('')
                self.label_partno_page4[x].setText('')
                self.label_partname_page4[x].setText('')
                self.label_barcode_page4[x].setText('')
                self.label_satus_page4[x].setText('')
                self.jiechu_button_page4[x].setText('')
            self.label_88.setText('')  # 88是显示暂无记录那个标签
            self.label_88.repaint()
            self.label_220.setText('1')
            self.label_220.repaint()
            # self.guazhang=(('LCH-2', '4X90S91831', 'CABLE USB-C TO ETHERNET ADP SC10P27625', '2022012111770'),('LCH-1', '20Y6Z9EBCN', 'THINKPAD X1 EXTREME GEN 4 SVT19 PRC', '2022012621453'))
            for each in self.guazhang_result_page4:
                if self.lineEdit_8.text() in each[2]:  # 如果搜索的名字被包含在PARTname里
                    page4_chaxun_partname_list.append(each)
            self.guazhang_result_page4 = tuple(page4_chaxun_partname_list)
            self.zongyeshu, self.yushu = len(self.guazhang_result_page4) // 13, len(self.guazhang_result_page4) % 13  # //取整，%取余数, 后面会用
            if self.zongyeshu != 0 and self.yushu != 0:
                self.zongyeshu = self.zongyeshu + 1
            if self.zongyeshu == 0:
                self.label_228.setText('/1')
                self.label_228.repaint()
            else:
                self.label_228.setText(f'/{self.zongyeshu}')
                self.label_228.repaint()
            if len(self.guazhang_result_page4) == 0:
                self.label_88.setText('暂无记录')
                self.label_88.repaint()
                self.label_112.setText('0')
                self.label_112.repaint()
            else:
                self.label_112.setText(str(len(self.guazhang_result_page4)))
                self.label_112.repaint()
                if len(self.guazhang_result_page4) <= 13:
                    for i in range(0, len(self.guazhang_result_page4)):
                        self.label_projectname_page4[i].setText(' ' + self.guazhang_result_page4[i][0])
                        self.label_partno_page4[i].setText('  ' + self.guazhang_result_page4[i][1])
                        self.label_partname_page4[i].setText(' ' + self.guazhang_result_page4[i][2])
                        self.label_barcode_page4[i].setText(self.guazhang_result_page4[i][3])
                        self.label_satus_page4[i].setText('   ' + self.guazhang_result_page4[i][4])
                        self.jiechu_button_page4[i].setText('                                     借出')
                        if self.guazhang_result_page4[i][4] == '已借出':
                            self.jiechu_button_page4[i].setEnabled(False)
                        else:
                            self.jiechu_button_page4[i].setEnabled(True)
                else:  # 大于13
                    for i in range(0, 13):
                        self.label_projectname_page4[i].setText(' ' + self.guazhang_result_page4[i][0])
                        self.label_partno_page4[i].setText('  ' + self.guazhang_result_page4[i][1])
                        self.label_partname_page4[i].setText(' ' + self.guazhang_result_page4[i][2])
                        self.label_barcode_page4[i].setText(self.guazhang_result_page4[i][3])
                        self.label_satus_page4[i].setText('   ' + self.guazhang_result_page4[i][4])
                        self.jiechu_button_page4[i].setText('                                     借出')
                        if self.guazhang_result_page4[i][4] == '已借出':
                            self.jiechu_button_page4[i].setEnabled(False)
                        else:
                            self.jiechu_button_page4[i].setEnabled(True)
        elif self.lineEdit_7.text() != '' and self.lineEdit_8.text() != '':  # part 和 project name不为空
            page4_chaxun_part_project_name_list = []  # 准备一个新的列表，筛选条件后，插入新的元素，依赖前面生成的所有的总的返回结果self.guazhang_result_page4，因为元组不能插入元素，写写列表后面转元组
            for x in range(0, 13):
                self.label_projectname_page4[x].setText('')
                self.label_partno_page4[x].setText('')
                self.label_partname_page4[x].setText('')
                self.label_barcode_page4[x].setText('')
                self.label_satus_page4[x].setText('')
                self.jiechu_button_page4[x].setText('')
            self.label_88.setText('')  # 88是显示暂无记录那个标签
            self.label_88.repaint()
            self.label_220.setText('1')
            self.label_220.repaint()
            # self.guazhang=(('LCH-2', '4X90S91831', 'CABLE USB-C TO ETHERNET ADP SC10P27625', '2022012111770'),('LCH-1', '20Y6Z9EBCN', 'THINKPAD X1 EXTREME GEN 4 SVT19 PRC', '2022012621453'))
            for each in self.guazhang_result_page4:
                if self.lineEdit_8.text() in each[2] and self.lineEdit_7.text() in each[0]:  # 如果搜索的名字被包含在PARTname里
                    page4_chaxun_part_project_name_list.append(each)
            self.guazhang_result_page4 = tuple(page4_chaxun_part_project_name_list)
            self.zongyeshu, self.yushu = len(self.guazhang_result_page4) // 13, len(self.guazhang_result_page4) % 13  # //取整，%取余数, 后面会用
            if self.zongyeshu != 0 and self.yushu != 0:
                self.zongyeshu = self.zongyeshu + 1
            if self.zongyeshu == 0:
                self.label_228.setText('/1')
                self.label_228.repaint()
            else:
                self.label_228.setText(f'/{self.zongyeshu}')
                self.label_228.repaint()
            if len(self.guazhang_result_page4) == 0:
                self.label_88.setText('暂无记录')
                self.label_88.repaint()
                self.label_112.setText('0')
                self.label_112.repaint()
            else:
                self.label_112.setText(str(len(self.guazhang_result_page4)))
                self.label_112.repaint()
                if len(self.guazhang_result_page4) <= 13:
                    for i in range(0, len(self.guazhang_result_page4)):
                        self.label_projectname_page4[i].setText(' ' + self.guazhang_result_page4[i][0])
                        self.label_partno_page4[i].setText('  ' + self.guazhang_result_page4[i][1])
                        self.label_partname_page4[i].setText(' ' + self.guazhang_result_page4[i][2])
                        self.label_barcode_page4[i].setText(self.guazhang_result_page4[i][3])
                        self.label_satus_page4[i].setText('   ' + self.guazhang_result_page4[i][4])
                        self.jiechu_button_page4[i].setText('                                     借出')
                        if self.guazhang_result_page4[i][4] == '已借出':
                            self.jiechu_button_page4[i].setEnabled(False)
                        else:
                            self.jiechu_button_page4[i].setEnabled(True)
                else:  # 大于13
                    for i in range(0, 13):
                        self.label_projectname_page4[i].setText(' ' + self.guazhang_result_page4[i][0])
                        self.label_partno_page4[i].setText('  ' + self.guazhang_result_page4[i][1])
                        self.label_partname_page4[i].setText(' ' + self.guazhang_result_page4[i][2])
                        self.label_barcode_page4[i].setText(self.guazhang_result_page4[i][3])
                        self.label_satus_page4[i].setText('   ' + self.guazhang_result_page4[i][4])
                        self.jiechu_button_page4[i].setText('                                     借出')
                        if self.guazhang_result_page4[i][4] == '已借出':
                            self.jiechu_button_page4[i].setEnabled(False)
                        else:
                            self.jiechu_button_page4[i].setEnabled(True)
        else:#如果2个都为空, 就什么都不做
            pass
        list1_page4=[self.label_92.text(),self.label_120.text(),self.label_121.text(),self.label_122.text(),self.label_123.text()]#第一行对应的label的各项内容project name 等
        list2_page4 = [self.label_93.text(), self.label_130.text(), self.label_129.text(), self.label_131.text(), self.label_127.text()]
        list3_page4 = [self.label_94.text(), self.label_137.text(), self.label_136.text(), self.label_138.text(), self.label_134.text()]  # 第一行对应的label的各项内容project name 等
        list4_page4 = [self.label_95.text(), self.label_145.text(), self.label_143.text(), self.label_141.text(), self.label_144.text()]
        list5_page4 = [self.label_98.text(), self.label_148.text(), self.label_150.text(), self.label_149.text(), self.label_153.text()]
        list6_page4 = [self.label_99.text(), self.label_156.text(), self.label_155.text(), self.label_160.text(), self.label_157.text()]
        list7_page4 = [self.label_166.text(), self.label_168.text(), self.label_162.text(), self.label_164.text(), self.label_165.text()]
        list8_page4 = [self.label_175.text(), self.label_177.text(), self.label_173.text(), self.label_170.text(), self.label_171.text()]
        list9_page4 = [self.label_178.text(), self.label_179.text(), self.label_181.text(), self.label_184.text(), self.label_185.text()]
        list10_page4 = [self.label_188.text(), self.label_187.text(), self.label_186.text(), self.label_191.text(), self.label_192.text()]
        list11_page4 = [self.label_194.text(), self.label_199.text(), self.label_197.text(), self.label_196.text(), self.label_200.text()]
        list12_page4 = [self.label_209.text(), self.label_208.text(), self.label_204.text(), self.label_207.text(), self.label_203.text()]
        list13_page4 = [self.label_226.text(), self.label_227.text(), self.label_214.text(), self.label_225.text(), self.label_213.text()]
        self.jiechu_button_duiying_label_page4={'pushButton_24':list1_page4,'pushButton_47':list2_page4,'pushButton_26':list3_page4,'pushButton_48':list4_page4,'pushButton_27':list5_page4,'pushButton_49':list6_page4,'pushButton_29':list7_page4,'pushButton_50':list8_page4,'pushButton_30':list9_page4,'pushButton_51':list10_page4,'pushButton_31':list11_page4,'pushButton_52':list12_page4,'pushButton_33':list13_page4}
    def show_frame3_page4_button24(self):#点击借出的时候显示借出框
        self.frame_3.show()
        self.lineEdit_9.clear()
        self.lineEdit_10.clear()
        self.label_97.setText('')
        self.label_97.setText('pushButton_24')
        self.label_96.setText('')
    def show_frame3_page4_button26(self):#点击借出的时候显示借出框
        self.frame_3.show()
        self.lineEdit_9.clear()
        self.lineEdit_10.clear()
        self.label_97.setText('')
        self.label_97.setText('pushButton_26')
        self.label_96.setText('')
    def show_frame3_page4_button47(self):  # 点击借出的时候显示借出框
        self.frame_3.show()
        self.lineEdit_9.clear()
        self.lineEdit_10.clear()
        self.label_97.setText('')
        self.label_97.setText('pushButton_47')
        self.label_96.setText('')
    def show_frame3_page4_button48(self):  # 点击借出的时候显示借出框
        self.frame_3.show()
        self.lineEdit_9.clear()
        self.lineEdit_10.clear()
        self.label_97.setText('')
        self.label_97.setText('pushButton_48')
        self.label_96.setText('')
    def show_frame3_page4_button49(self):  # 点击借出的时候显示借出框
        self.frame_3.show()
        self.lineEdit_9.clear()
        self.lineEdit_10.clear()
        self.label_97.setText('')
        self.label_97.setText('pushButton_49')
        self.label_96.setText('')
    def show_frame3_page4_button50(self):  # 点击借出的时候显示借出框
        self.frame_3.show()
        self.lineEdit_9.clear()
        self.lineEdit_10.clear()
        self.label_97.setText('')
        self.label_97.setText('pushButton_50')
        self.label_96.setText('')
    def show_frame3_page4_button51(self):  # 点击借出的时候显示借出框
        self.frame_3.show()
        self.lineEdit_9.clear()
        self.lineEdit_10.clear()
        self.label_97.setText('')
        self.label_97.setText('pushButton_51')
        self.label_96.setText('')
    def show_frame3_page4_button52(self):  # 点击借出的时候显示借出框
        self.frame_3.show()
        self.lineEdit_9.clear()
        self.lineEdit_10.clear()
        self.label_97.setText('')
        self.label_97.setText('pushButton_52')
        self.label_96.setText('')
    def show_frame3_page4_button27(self):  # 点击借出的时候显示借出框
        self.frame_3.show()
        self.lineEdit_9.clear()
        self.lineEdit_10.clear()
        self.label_97.setText('')
        self.label_97.setText('pushButton_27')
        self.label_96.setText('')
    def show_frame3_page4_button29(self):  # 点击借出的时候显示借出框
        self.frame_3.show()
        self.lineEdit_9.clear()
        self.lineEdit_10.clear()
        self.label_97.setText('')
        self.label_97.setText('pushButton_29')
        self.label_96.setText('')
    def show_frame3_page4_button30(self):  # 点击借出的时候显示借出框
        self.frame_3.show()
        self.lineEdit_9.clear()
        self.lineEdit_10.clear()
        self.label_97.setText('')
        self.label_97.setText('pushButton_30')
        self.label_96.setText('')
    def show_frame3_page4_button31(self):  # 点击借出的时候显示借出框
        self.frame_3.show()
        self.lineEdit_9.clear()
        self.lineEdit_10.clear()
        self.label_97.setText('')
        self.label_97.setText('pushButton_31')
        self.label_96.setText('')
    def show_frame3_page4_button33(self):  # 点击借出的时候显示借出框
        self.frame_3.show()
        self.lineEdit_9.clear()
        self.lineEdit_10.clear()
        self.label_97.setText('')
        self.label_97.setText('pushButton_33')
        self.label_96.setText('')
    def hide_frame3(self):
        self.frame_3.hide()
    def page4_jiqi_jiechu_queren(self):# 97 label是记录button编号的label标签，颜色和背景一样看不出来哈哈
        jiechuqingdan_table = 'jiechuqingdan_' + user#数据库表格
        self.label_96.setText('')
        if len(self.lineEdit_10.text())!=9 and self.lineEdit_9.text()!='':
            self.label_96.setText('请输入9位数工号')
        elif len(self.lineEdit_10.text())==9  and self.lineEdit_9.text()=='':
            self.label_96.setText('请输入借用人姓名')
        elif len(self.lineEdit_10.text())!=9 and self.lineEdit_9.text()=='':
            self.label_96.setText('请输入借用人姓名')
        else:
            try:
                db = pymysql.connect(host="localhost", user='root', password='jiang111', port=3306, database='lily', connect_timeout=4)
                cusor = db.cursor()
                try:
                    sql=f"UPDATE {self.table_str} set Status='已借出' WHERE Barcode='{self.jiechu_button_duiying_label_page4[self.label_97.text()][3]}'"#数据库挂账表
                    project_name=self.jiechu_button_duiying_label_page4[self.label_97.text()][0]
                    part_no=self.jiechu_button_duiying_label_page4[self.label_97.text()][1]
                    part_name = self.jiechu_button_duiying_label_page4[self.label_97.text()][2]
                    barcode = self.jiechu_button_duiying_label_page4[self.label_97.text()][3]
                    jiechu_shijian=time.strftime('%Y-%m-%d', time.localtime())#时间的字符串格式 获取当前本地时间，这里也是当作借入时间
                    sq1_1=f"INSERT INTO {jiechuqingdan_table} values ('{project_name}','{part_no}','{part_name}','{barcode}','{jiechu_shijian}','{self.lineEdit_9.text()}','{self.lineEdit_10.text()}')"
                    sql_3=f"UPDATE {self.wuzichazhao} set jieyong_ren='{self.lineEdit_9.text()}' WHERE Barcode='{self.jiechu_button_duiying_label_page4[self.label_97.text()][3]}'"#数据库挂账表
                    jieruqingdan_table = 'jieruqingdan_' +  self.lineEdit_10.text()# 数据库表格
                    cusor.execute(sql)
                    cusor.execute(sq1_1)
                    cusor.execute(sql_3)
                    if self.lineEdit_10.text() in gonghao_all:
                        sq1_2=f"INSERT INTO {jieruqingdan_table} values ('{project_name}','{part_no}','{part_name}','{barcode}','{jiechu_shijian}','{user_to_name[user]}')"
                        cusor.execute(sq1_2)
                    db.commit()
                    self.frame_3.hide()
                except:
                    db.rollback()
                    self.label_96.setText('借出失败')
            except:
                self.label_96.setText('借出失败')
    #上面是page4里包含的函数

    # 下面是page5里包含的函数
    def page5(self):
        self.stackedWidget.setCurrentIndex(4)
        self.label_113.setText('')
    def xuanzhe_ruku_file(self):
        self.my_file_path = QFileDialog.getOpenFileName(None, '选择入库文件', '*.xlsx')#返回('C:/Users/7/Downloads/RD各部門庫存掛賬統計與管控0409.xlsx', 'All Files (*)')
    def xuanzhe_chuku_file(self):
        self.my_file_path = QFileDialog.getOpenFileName(None, '选择出库文件', '*.xlsx')#返回('C:/Users/7/Downloads/RD各部門庫存掛賬統計與管控0409.xlsx', 'All Files (*)')
    def ruku(self):
        #self.panduan_user()
        try:
            self.label_113.setText('~~~~~正在入库,请稍等~~~~~')
            self.label_113.repaint()
            db = pymysql.connect(host="localhost", user='root', password='jiang111', port=3306, database='lily', connect_timeout=3)
            cusor = db.cursor()
            excel_ruku = openpyxl.load_workbook(self.my_file_path[0])  # self.my_file_path[0]就是文件路径
            sheet_1 = excel_ruku['Sheet1']
            sheet_1_max_rows = sheet_1.max_row
            for every_row in range(2, sheet_1_max_rows + 1):  # 这里是匹配sheet1的最大行数，除去第一行表头，+1是因为range(2,9）行是循环2到8行.所以加1刚刚好
                try:
                    every_row_list = []
                    for cell in sheet_1[every_row]:  # sheet[i] i是就代表sheet的第几行，cell是代表这一行中的每一个单元格
                        every_row_list.append(cell.value)
                    table_str = 'guazhang_' + user
                    insert_sql = f"insert into {table_str} (Project_Name, Part_No, Part_Name,Barcode) values('{every_row_list[0]}','{every_row_list[1]}','{every_row_list[2]}','{every_row_list[3]}')"
                    cusor.execute(insert_sql)
                    insert_sql_1=f"insert into {self.wuzichazhao} (Project_Name,Part_Name,Barcode,guazhang_ren) values('{every_row_list[0]}','{every_row_list[2]}','{every_row_list[3]}','{user_to_name[user]}')"
                    cusor.execute(insert_sql_1)
                    db.commit()
                    sheet_1.cell(column=5, row=every_row, value='入库成功')
                    excel_ruku.save(self.my_file_path[0])
                except Exception as e:
                    db.rollback()
                    sheet_1.cell(column=5, row=every_row, value=str(e))
                    excel_ruku.save(self.my_file_path[0])
            self.label_113.setText("操作完成,详情请查看excel!!!")
            db.close()
        except:
            self.label_113.setText('入库失败,请检查网络或excel格式!!!')
    def tuiku(self):
        #self.panduan_user()
        try:
            self.label_113.setText('~~~~~正在退库,请稍等~~~~~')
            self.label_113.repaint()
            db = pymysql.connect(host="localhost", user='root', password='jiang111', port=3306, database='lily', connect_timeout=3)
            cusor = db.cursor()
            excel_ruku = openpyxl.load_workbook(self.my_file_path[0])  # self.my_file_path[0]就是文件路径
            sheet_1 = excel_ruku['Sheet1']
            sheet_1_max_rows = sheet_1.max_row
            for every_row in range(2, sheet_1_max_rows + 1):  # 这里是匹配sheet1的最大行数，除去第一行表头，+1是因为range(2,9）行是循环2到8行.所以加1刚刚好
                try:
                    every_row_list = []
                    for cell in sheet_1[every_row]:  # sheet[i] i是就代表sheet的第几行，cell是代表这一行中的每一个单元格
                        every_row_list.append(cell.value)
                    table_str='guazhang_'+user
                    select_sql = f"select * from {table_str} where Barcode='{every_row_list[0]}'"
                    cusor.execute(select_sql)
                    result_return = cusor.fetchone()
                    if result_return == None:  # 代表没有找到此设备
                        sheet_1.cell(column=2, row=every_row, value='此料号不存在')
                        excel_ruku.save(self.my_file_path[0])
                    else:
                        delete_sql = f"delete from {table_str} where Barcode='{every_row_list[0]}'"
                        delete_sql_1 = f"delete from {self.wuzichazhao} where Barcode='{every_row_list[0]}'"
                        cusor.execute(delete_sql)
                        cusor.execute(delete_sql_1)
                        db.commit()
                        sheet_1.cell(column=2, row=every_row, value='退库成功')
                        excel_ruku.save(self.my_file_path[0])
                except Exception as e:
                    db.rollback()
                    sheet_1.cell(column=2, row=every_row, value=str(e))
                    excel_ruku.save(self.my_file_path[0])
            self.label_113.setText("操作完成,详情请查看excel!!!")
            db.close()
        except:
            self.label_113.setText('退库失败,请检查网络或excel格式!!!')
    # 上面是page5里包含的函数

    def showTime(self):
        now=datetime.now()
        date_dict={0:'周一',1:'周二',2:'周三',3:'周四',4:'周五',5:'周六',6:'周日'}
        # 获取系统当前时间
        time = QDateTime.currentDateTime()
        # 设置系统时间的显示格式
        timeDisplay = time.toString('yyyy-MM-dd hh:mm:ss')
        # 在标签上显示时间
        weekday = now.weekday()# 获取星期几，返回0-6的整数（0代表周一，6代表周日）
        self.label_2.setText(timeDisplay[0:10]+'  '+date_dict[weekday]+''+timeDisplay[10:])
#------------------------------主界面---------------------------------------#

if __name__ == '__main__':
    QtCore.QCoreApplication.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling)  # 自适应分辨率
    app = QtWidgets.QApplication(sys.argv)
    win_login=loginwindow()
    win_login.show()
    win_main_interface = interfacewindow()
    sys.exit(app.exec())